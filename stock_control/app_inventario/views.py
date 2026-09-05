# views.py (consolidado y corregido)
import io
import time
from django.utils import timezone
import os
import json
import sqlite3
import logging
import pandas as pd

from io import BytesIO
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.utils.dateparse import parse_datetime
from datetime import datetime, time


from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Max, Q, OuterRef, Subquery
from django.http import JsonResponse, FileResponse, Http404
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.views.decorators.cache import cache_page
from django.views.decorators.csrf import csrf_exempt
from django.core.management import call_command

from app_inventario.services.printing import print_stock_total
from datetime import datetime, time, timedelta
from django.db.models import Sum, Count, Q, F, Avg, IntegerField
from django.db.models.functions import Cast



from .models import (
    BocaSalida, OrigenIngreso, ProductoFijo,
    RegistroMovimiento, GrupoMovimiento, StockBalde, ConciliacionBoca,
    ConfiguracionSistema,
)

_CONFIG_DEFAULTS = {
    'precio_cat1':                  '12500',
    'precio_cat2':                  '13500',
    'precio_cat3':                  '14500',
    'precio_gastronomico':          '15000',
    'precio_gastronomico_pistacho': '18000',
    'valor_inventario_kg':          '22500',
}

def get_config(clave, default=None):
    try:
        return ConfiguracionSistema.objects.get(clave=clave).valor
    except ConfiguracionSistema.DoesNotExist:
        return default if default is not None else _CONFIG_DEFAULTS.get(clave, '')
from app_inventario import models

logger = logging.getLogger(__name__)

# =========================================================
# Helpers
# =========================================================

def conectar_bd():
    """Conectar a la base de datos SQLite (utilidad puntual)."""
    conn = sqlite3.connect("inventario.db")
    conn.row_factory = sqlite3.Row
    return conn

def api_stock_detallado(request):
    productos = ProductoFijo.objects.annotate(
        stock_actual=Count('stockbalde', filter=Q(stockbalde__is_activo=True))
    )
    data = [
        {"nombre": p.nombre, "stock_minimo": p.stock_minimo, "cantidad": p.stock_actual}
        for p in productos
    ]
    return JsonResponse({"stock_detallado": data})

# --- Legacy: agregar_productos (agrega directo al stock SIN grupo/movimiento)
from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def agregar_productos(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body or "{}")
        productos = data.get("productos", [])
        if not productos:
            return JsonResponse({"error": "No se enviaron productos"}, status=400)

        for producto in productos:
            plu = producto.get("plu")
            peso = producto.get("peso")
            if not plu or peso is None:
                return JsonResponse({"error": "Datos incompletos"}, status=400)
            producto_obj = ProductoFijo.objects.get(plu=plu)
            StockBalde.objects.create(producto=producto_obj, peso=float(peso))

        return JsonResponse({"message": "Productos agregados exitosamente"}, status=200)
    except ProductoFijo.DoesNotExist:
        return JsonResponse({"error": "Producto no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# --- API CRUD PRODUCTOS

@csrf_exempt
def api_listar_productos(request):
    if request.method != "GET":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    productos = ProductoFijo.objects.all().order_by("nombre")

    data = [
        {
            "plu":          p.plu,
            "nombre":       p.nombre,
            "stock_minimo": p.stock_minimo,
            "is_activo":    p.is_activo,
        }
        for p in productos
    ]

    return JsonResponse({"productos": data})


@csrf_exempt
def api_crear_producto(request):
    try:
        data = json.loads(request.body or "{}")
    except:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    nombre = (data.get("nombre") or "").strip()
    plu = (data.get("plu") or "").strip().zfill(3)
    minimo = int(data.get("stock_minimo") or 0)

    if not nombre:
        return JsonResponse({"error": "El nombre es obligatorio"}, status=400)

    if ProductoFijo.objects.filter(plu=plu).exists():
        return JsonResponse({"error": f"Ya existe un producto con PLU {plu}"}, status=400)

    ProductoFijo.objects.create(nombre=nombre, plu=plu, stock_minimo=minimo)

    return JsonResponse({"success": True, "message": "Producto creado correctamente"})


@csrf_exempt
def api_eliminar_producto(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON inválido"}, status=400)

    plu = (data.get("plu") or "").strip()
    if not plu:
        return JsonResponse({"success": False, "error": "Falta 'plu' del producto"}, status=400)

    try:
        p = ProductoFijo.objects.get(plu=plu)
    except ProductoFijo.DoesNotExist:
        return JsonResponse({"success": False, "error": f"Producto con PLU {plu} no encontrado"}, status=404)

    # (Opcional) Evitar borrar si tiene stock/movimientos asociados
    tiene_stock = StockBalde.objects.filter(producto=p).exists()
    tiene_movs = RegistroMovimiento.objects.filter(producto=p).exists()
    if tiene_stock or tiene_movs:
        return JsonResponse({
            "success": False,
            "error": "No se puede eliminar el producto porque tiene stock o movimientos asociados."
        }, status=400)

    p.delete()
    return JsonResponse({"success": True, "message": f"Producto {p.nombre} (PLU {plu}) eliminado correctamente."})


@csrf_exempt
def api_actualizar_producto(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Método no permitido"},
            status=405
        )

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse(
            {"success": False, "error": "JSON inválido"},
            status=400
        )

    plu = (data.get("plu") or "").strip()
    nombre = (data.get("nombre") or "").strip()
    stock_minimo = data.get("stock_minimo")

    if not plu:
        return JsonResponse(
            {"success": False, "error": "Falta PLU del producto"},
            status=400
        )

    try:
        p = ProductoFijo.objects.get(plu=plu)
    except ProductoFijo.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": f"No existe producto con PLU {plu}"},
            status=404
        )

    # Actualizar nombre (si vino)
    if nombre:
        p.nombre = nombre

    # Actualizar stock_minimo (si vino)
    if stock_minimo not in (None, ""):
        try:
            p.stock_minimo = int(stock_minimo)
        except ValueError:
            return JsonResponse(
                {"success": False, "error": "stock_minimo debe ser numérico"},
                status=400
            )

    p.save()

    return JsonResponse(
        {"success": True, "message": "Producto actualizado correctamente"},
        status=200
    )

@csrf_exempt
def api_toggle_plu_activo(request):
    """POST {"plu": "001"} — activa/desactiva un PLU.
    PLUs inactivos no aparecen en la impresión de stock."""
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    plu = (data.get("plu") or "").strip()
    if not plu:
        return JsonResponse({"error": "Falta 'plu'"}, status=400)

    try:
        p = ProductoFijo.objects.get(plu=plu)
    except ProductoFijo.DoesNotExist:
        return JsonResponse({"error": f"PLU {plu} no encontrado"}, status=404)

    p.is_activo = not p.is_activo
    p.save(update_fields=["is_activo"])
    estado = "activado" if p.is_activo else "desactivado"
    return JsonResponse({
        "success": True,
        "is_activo": p.is_activo,
        "plu": plu,
        "message": f"PLU {plu} ({p.nombre}) {estado} correctamente.",
    })

# ---  FIN API CRUD PRODUCTOS

def actualizar_stock_minimo(request):
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON inválido"}, status=400)

    productos = data.get("productos", [])
    if not isinstance(productos, list):
        return JsonResponse({"success": False, "error": "'productos' debe ser una lista"}, status=400)

    actualizados = 0
    for producto in productos:                       # <— usá SIEMPRE el mismo nombre
        plu = producto.get("plu")
        minimo = producto.get("stock_minimo")
        if plu is None or minimo is None:
            continue
        actualizados += (
            ProductoFijo.objects
            .filter(plu=plu)
            .update(stock_minimo=minimo)
        )

    return JsonResponse({"success": True, "message": "OK", "actualizados": actualizados}, status=200)


def _actualizar_total_grupo(grupo_id, tipo, origen=None, destino_nombre=None):
    agg = (RegistroMovimiento.objects
           .filter(grupo_id=grupo_id)
           .aggregate(total=Sum('peso'), cantidad=Count('id')))
    total = agg['total'] or 0
    cant = agg['cantidad'] or 0

    destino_obj = None
    if destino_nombre:
        destino_obj = BocaSalida.objects.filter(nombre=destino_nombre).first()

    GrupoMovimiento.objects.update_or_create(
        grupo_id=grupo_id,
        defaults={
            'tipo': tipo,
            'origen': origen if tipo in ('ingreso', 'devolucion') else None,
            'destino': destino_obj if tipo in ('salida', 'retiro') else None,
            'total_peso': total,
            'cantidad_items': cant,
        }
    )




def _print_group_if_enabled(grupo_id: int):
    if not getattr(settings, "PRINT_FROM_VIEWS", False):
        return
    try:
        from .services.printing import print_grupo_movimiento
        copias = getattr(settings, "PRINT_COPIAS", 1)
        print_grupo_movimiento(grupo_id, copias=copias)
        logger.info("Ticket de grupo #%s impreso (desde view)", grupo_id)
    except Exception:
        logger.exception("Error al imprimir ticket del grupo #%s (desde view)", grupo_id)

@csrf_exempt
def reimprimir_ticket(request, grupo_id: int):
    """
    Reimprime el comprobante del movimiento agrupado (grupo_id).
    Acepta 'copias' opcional por POST o ?copias=
    """
    try:
        copias = int(
            request.POST.get("copias")
            or request.GET.get("copias")
            or getattr(settings, "PRINT_COPIAS", 1)
            or 1
        )
        from .services.printing import print_grupo_movimiento
        print_grupo_movimiento(grupo_id, copias=copias)
        logger.info("Reimpresión solicitada para grupo #%s (%s copias)", grupo_id, copias)
        return JsonResponse({"ok": True, "message": f"Reimpresión enviada para el grupo #{grupo_id} ({copias} copia/s)."})
    except Exception as e:
        logger.exception("Error reimprimiendo grupo #%s", grupo_id)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

@csrf_exempt
def imprimir_stock_total(request):
    """
    Endpoint para imprimir el stock total en la impresora de tickets.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método no permitido"}, status=405)

    try:
        print_stock_total()
        return JsonResponse({"ok": True, "message": "Impresión de stock enviada a la impresora."})
    except Exception as e:
        # si querés loggear:
        # logger.exception("Error imprimiendo stock total")
        return JsonResponse({"ok": False, "error": str(e)}, status=500)
# =========================================================
# Vistas principales
# =========================================================

def _categorizar_producto(plu):
    """Clasifica un producto por rango de PLU en una de 3 categorías."""
    try:
        plu_int = int(plu)
    except (ValueError, TypeError):
        return 'helado'
    if plu_int >= 100:
        return 'gastronomico'
    if 89 <= plu_int <= 98:
        return 'barra_torta'
    return 'helado'


def _plu_listas_categoria():
    """Devuelve (plu_helados, plu_tortas, plu_gastro) como listas de str PKs.
    Cada lista contiene los valores de plu (CharField) del rango correspondiente:
      helados      PLU 1-88
      tortas/barras PLU 89-98
      gastronómico  PLU 100-199
    Usado como filtro en anotaciones condicionales Sum."""
    helados, tortas, gastro = [], [], []
    for plu_str in ProductoFijo.objects.values_list("plu", flat=True):
        try:
            v = int(plu_str)
        except (ValueError, TypeError):
            continue
        if 1 <= v <= 88:
            helados.append(plu_str)
        elif 89 <= v <= 98:
            tortas.append(plu_str)
        elif 100 <= v <= 199:
            gastro.append(plu_str)
    return helados, tortas, gastro


def index(request):
    stock_resumido = ProductoFijo.objects.annotate(
        cantidad=Count('stockbalde', filter=Q(stockbalde__is_activo=True))
    )
    tot_balde = StockBalde.objects.filter(is_activo=True).count()
    tot_kilos = StockBalde.objects.filter(is_activo=True).aggregate(s=Sum('peso'))['s'] or 0

    cats = {
        'helado':       {'baldes': 0, 'kilos': 0.0},
        'barra_torta':  {'baldes': 0, 'kilos': 0.0},
        'gastronomico': {'baldes': 0, 'kilos': 0.0},
    }
    productos = ProductoFijo.objects.annotate(
        cantidad=Count('stockbalde', filter=Q(stockbalde__is_activo=True)),
        kg_total=Sum('stockbalde__peso', filter=Q(stockbalde__is_activo=True)),
    ).values('plu', 'cantidad', 'kg_total')
    for p in productos:
        cat = _categorizar_producto(p['plu'])
        cats[cat]['baldes'] += p['cantidad'] or 0
        cats[cat]['kilos'] += float(p['kg_total'] or 0)
    for cat in cats:
        cats[cat]['kilos'] = round(cats[cat]['kilos'], 2)

    from version import APP_VERSION
    return render(request, "index.html", {
        "stock_resumido": stock_resumido,
        "total_baldes": tot_balde,
        "total_kilos": round(float(tot_kilos), 2),
        "cats": cats,
        "APP_VERSION": APP_VERSION,
    })


def cargar_productos_desde_excel():
    """
    Carga productos desde productos.xlsx (Nombre, PLU) ubicado junto a la app.
    """
    archivo_excel = os.path.join(os.path.dirname(__file__), "productos.xlsx")
    try:
        df = pd.read_excel(archivo_excel)
        if 'Nombre' not in df.columns or 'PLU' not in df.columns:
            return {"error": "El archivo Excel no tiene las columnas necesarias (Nombre, PLU)"}

        for _, row in df.iterrows():
            nombre = row['Nombre']
            plu = str(row['PLU']).zfill(3)
            ProductoFijo.objects.get_or_create(plu=plu, nombre=nombre)

        return {"message": "Productos cargados exitosamente"}

    except Exception as e:
        return {"error": f"Error procesando el archivo: {str(e)}"}


@csrf_exempt
def cargar_productos_excel(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        fs = FileSystemStorage(location="uploads/")
        nombre_archivo = fs.save(archivo.name, archivo)
        ruta_archivo = fs.path(nombre_archivo)

        try:
            df = pd.read_excel(ruta_archivo)
            if "Nombre" not in df.columns or "PLU" not in df.columns:
                return JsonResponse({"error": "El archivo debe contener las columnas 'Nombre' y 'PLU'"}, status=400)

            for _, row in df.iterrows():
                nombre = row["Nombre"]
                plu = str(row["PLU"]).zfill(3)
                ProductoFijo.objects.get_or_create(plu=plu, nombre=nombre)

            return JsonResponse({"message": "Productos cargados correctamente"})

        except Exception as e:
            return JsonResponse({"error": f"Error al procesar el archivo: {str(e)}"}, status=500)

    return JsonResponse({"error": "No se envió ningún archivo"}, status=400)


def importar_productos(request):
    resultado = cargar_productos_desde_excel()
    return JsonResponse(resultado)


def exportar_productos_excel(request):
    """
    Exporta un Excel con las columnas:
    - Nombre
    - PLU

    Lo podés usar como plantilla: agregás nuevas filas y luego
    lo volvés a subir por /cargar_excel/ para alta masiva.
    """
    productos = ProductoFijo.objects.all().order_by("plu")

    filas = [
        {"Nombre": p.nombre, "PLU": p.plu}
        for p in productos
    ]

    # Si no hay productos aún, devolvemos solo el header
    if filas:
        df = pd.DataFrame(filas)
    else:
        df = pd.DataFrame(columns=["Nombre", "PLU"])

    # Escribir a un buffer en memoria
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Productos")

    buffer.seek(0)

    filename = f"productos_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



@cache_page(5)
def obtener_stock(request):
    try:
        productos_stock = ProductoFijo.objects.annotate(
            cantidad=Count("stockbalde", filter=Q(stockbalde__is_activo=True)),
            kg_total=Sum("stockbalde__peso", filter=Q(stockbalde__is_activo=True)),
        ).values("plu", "nombre", "cantidad", "stock_minimo", "kg_total")

        stock_list = []
        cats = {
            'helado':       {'baldes': 0, 'kilos': 0.0},
            'barra_torta':  {'baldes': 0, 'kilos': 0.0},
            'gastronomico': {'baldes': 0, 'kilos': 0.0},
        }
        for p in productos_stock:
            kg = round(float(p['kg_total'] or 0), 3)
            stock_list.append({
                'plu': p['plu'],
                'nombre': p['nombre'],
                'cantidad': p['cantidad'],
                'stock_minimo': p['stock_minimo'],
                'kg_total': kg,
            })
            cat = _categorizar_producto(p['plu'])
            cats[cat]['baldes'] += p['cantidad'] or 0
            cats[cat]['kilos'] += kg
        for cat in cats:
            cats[cat]['kilos'] = round(cats[cat]['kilos'], 2)

        return JsonResponse({"stock": stock_list, "categorias": cats})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def stock_detallado(request):
    stock = StockBalde.objects.select_related("producto").all()
    return render(request, "stock_detallado.html", {"stock_detallado": stock})


def historial(request):
    movimientos_list = RegistroMovimiento.objects.all().order_by('-timestamp')
    paginator = Paginator(movimientos_list, 10)
    page_number = request.GET.get("page", 1)

    movimientos = paginator.get_page(page_number)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = [
            {
                "grupo_id": mov.grupo_id,
                "tipo": mov.tipo,
                "timestamp": mov.timestamp.strftime("%d/%m/%Y %H:%M")
            }
            for mov in movimientos
        ]
        return JsonResponse({
            "movimientos": data,
            "has_next": movimientos.has_next(),
            "has_previous": movimientos.has_previous(),
            "current_page": movimientos.number,
            "num_pages": movimientos.paginator.num_pages,
        })

    return render(request, "historial_movimientos.html", {"movimientos": movimientos})


# importa tus modelos:
# from .models import RegistroMovimiento, GrupoMovimiento, StockBalde, ProductoFijo

# views.py (agrega/asegúrate de tener estos imports arriba del archivo)
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.db.models import Q, Sum, OuterRef, Subquery
from django.core.paginator import Paginator


from django.utils.timezone import make_aware

import datetime as dt
from django.utils.timezone import make_aware



def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y %H:%M")

def _build_periodo_label(dt_desde, dt_hasta) -> str:
    if dt_desde and dt_hasta:
        return f"{_fmt_dt(dt_desde)} — {_fmt_dt(dt_hasta)}"
    if dt_desde:
        return f"desde {_fmt_dt(dt_desde)}"
    if dt_hasta:
        return f"hasta {_fmt_dt(dt_hasta)}"
    return "sin filtro (todos)"


def _parse_dt_local(s: str | None, is_end: bool = False):
    """
    Acepta valores de <input type="date"> (YYYY-MM-DD)
    o <input type="datetime-local"> (YYYY-MM-DDTHH:MM).
    Devuelve datetime naive. Si sólo viene fecha:
      - is_end=False -> 00:00:00
      - is_end=True  -> 23:59:59.999999
    """
    if not s:
        return None
    s = s.strip()
    # Fix Bug #3: parse_datetime("2024-01-15") retorna datetime a medianoche aunque
    # no haya componente de hora, por lo que el bloque is_end nunca se alcanzaba.
    # Solución: solo usar parse_datetime si el string contiene hora (T o :).
    if "T" in s or ":" in s:
        dt = parse_datetime(s)
        if dt:
            return dt
    # Para fechas sin hora, aplicar inicio/fin del día según is_end.
    d = parse_date(s)
    if d:
        if is_end:
            return datetime.combine(d, time(23, 59, 59, 999999))
        return datetime.combine(d, time(0, 0, 0))
    return None


def historial_movimientos(request):
    """
    Historial + modo activos.
    Incluye:
      - Resumen de RETIROS por PLU (orden ?orden=kg_desc|kg_asc, export ?export=xlsx)
      - Resumen de INGRESOS por PLU (orden ?orden_ing=kg_desc|kg_asc, export ?export=ing_xlsx)
    Acepta 'desde' y 'hasta' con fecha u hora (datetime-local).
    """
    # -------- Flags y filtros ----------
    solo_activos = request.GET.get("solo_activos") in {"1", "true", "True"}

    desde_str = (request.GET.get("desde") or "").strip()
    hasta_str = (request.GET.get("hasta") or "").strip()
    local = (request.GET.get("local") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip().lower()
    gusto = (request.GET.get("gusto") or "").strip()
    codigo = (request.GET.get("codigo") or "").strip()

    plus_str = (request.GET.get("plus") or "").strip()
    plus_list = [p.strip() for p in plus_str.replace(";", ",").split(",") if p.strip()]

    # Ordenes y export
    orden = (request.GET.get("orden") or "kg_desc").lower()           # retiros
    orden_ing = (request.GET.get("orden_ing") or "kg_desc").lower()   # ingresos
    export = (request.GET.get("export") or "").lower()                # 'xlsx' | 'ing_xlsx'

    # Fechas / horas
    dt_desde = _parse_dt_local(desde_str, is_end=False)
    dt_hasta = _parse_dt_local(hasta_str, is_end=True)

    # ============================================================
    # MODO "SOLO ACTIVOS"
    # ============================================================
    if solo_activos:
        qs = (
            StockBalde.objects
            .filter(is_activo=True)
            .select_related("producto")
            .only("id", "codigo_barras", "peso", "producto__nombre", "producto__plu", "timestamp")
        )
        if dt_desde:
            qs = qs.filter(timestamp__gte=dt_desde)
        if dt_hasta:
            qs = qs.filter(timestamp__lte=dt_hasta)
        if gusto:
            qs = qs.filter(producto__nombre__icontains=gusto)
        if codigo:
            qs = qs.filter(codigo_barras__icontains=codigo)

        qs = qs.order_by("producto__nombre", "timestamp", "id")
        total_kg_global = qs.aggregate(s=Sum("peso"))["s"] or 0

        page_number = request.GET.get("page", 1)
        paginator = Paginator(qs, 20)
        activos_page = paginator.get_page(page_number)

        return render(
            request,
            "historial_movimientos.html",
            {
                "modo": "activos",
                "movimientos": activos_page,
                "filtros": {
                    "desde": desde_str,
                    "hasta": hasta_str,
                    "local": local,
                    "tipo": tipo,
                    "gusto": gusto,
                    "codigo": codigo,
                    "plus": plus_str,
                    "orden": orden,
                    "orden_ing": orden_ing,
                    "solo_activos": True,
                },
                "total_kg_global": total_kg_global,
                "totales_plus": None,
                "totales_plus_ing": None,
            },
        )

    # ========================================
    # MODO HISTORIAL (último de cada grupo_id)
    # ========================================
    base_qs = RegistroMovimiento.objects.all()

    if dt_desde:
        base_qs = base_qs.filter(timestamp__gte=dt_desde)
    if dt_hasta:
        base_qs = base_qs.filter(timestamp__lte=dt_hasta)

    if local:
        base_qs = base_qs.filter(
            Q(boca_salida__icontains=local) |
            Q(origen__icontains=local) |
            Q(destino__nombre__icontains=local)
        )

    if tipo:
        if tipo in ("retiro", "salida"):
            base_qs = base_qs.filter(tipo__in=["retiro", "salida"])
        elif tipo == "ingreso":
            base_qs = base_qs.filter(tipo="ingreso")
        elif tipo == "devolucion":
            base_qs = base_qs.filter(tipo="devolucion")

    if gusto:
        base_qs = base_qs.filter(producto__nombre__icontains=gusto)
    if codigo:
        base_qs = base_qs.filter(codigo_barras__icontains=codigo)

    # ------- grilla principal: último por grupo -------
    latest_in_group = (
        base_qs.filter(grupo_id=OuterRef("grupo_id"))
               .order_by("-timestamp", "-id")
               .values("id")[:1]
    )
    latest_ids_subq = (
        base_qs.values("grupo_id")
               .annotate(latest_id=Subquery(latest_in_group))
               .values("latest_id")
    )
    movimientos_qs = (
        RegistroMovimiento.objects
        .filter(id__in=Subquery(latest_ids_subq))
        .select_related("producto", "destino")
        .order_by("-timestamp", "-id")
    )
    total_kg_global = base_qs.aggregate(s=Sum("peso"))["s"] or 0

    page_number = request.GET.get("page", 1)
    paginator = Paginator(movimientos_qs, 20)
    movimientos_page = paginator.get_page(page_number)

    # ------------------------------------------------------------
    # Resumen de RETIROS por PLU
    # ------------------------------------------------------------
    qs_retiros = base_qs.filter(tipo__in=["retiro", "salida"])
    if plus_list:
        qs_retiros = qs_retiros.filter(producto__plu__in=plus_list)

    detalle_por_plu_qs = (
        qs_retiros
        .values("producto__plu", "producto__nombre")
        .annotate(cant=Count("id"), kg=Sum("peso"))
    )
    if orden == "kg_asc":
        detalle_por_plu_qs = detalle_por_plu_qs.order_by("kg", "producto__nombre", "producto__plu")
    else:
        detalle_por_plu_qs = detalle_por_plu_qs.order_by("-kg", "producto__nombre", "producto__plu")

    detalle_por_plu = list(detalle_por_plu_qs)
    totales_plus = None
    if detalle_por_plu:
        # Devoluciones por PLU (mismos filtros de fecha/gusto/código, sin filtro de tipo)
        qs_dev_plu = RegistroMovimiento.objects.filter(tipo='devolucion')
        if dt_desde:
            qs_dev_plu = qs_dev_plu.filter(timestamp__gte=dt_desde)
        if dt_hasta:
            qs_dev_plu = qs_dev_plu.filter(timestamp__lte=dt_hasta)
        if gusto:
            qs_dev_plu = qs_dev_plu.filter(producto__nombre__icontains=gusto)
        if codigo:
            qs_dev_plu = qs_dev_plu.filter(codigo_barras__icontains=codigo)
        if plus_list:
            qs_dev_plu = qs_dev_plu.filter(producto__plu__in=plus_list)
        dev_por_plu = {
            row['producto__plu']: float(row['kg_dev'] or 0)
            for row in qs_dev_plu.values('producto__plu').annotate(kg_dev=Sum('peso'))
        }

        detalle_enriquecido = []
        for r in detalle_por_plu:
            kg_ret = float(r['kg'] or 0)
            kg_dev = dev_por_plu.get(r['producto__plu'], 0.0)
            detalle_enriquecido.append({
                **r,
                'kg_devuelto': round(kg_dev, 3),
                'kg_neto': round(kg_ret - kg_dev, 3),
            })

        total_kg_ret = float(sum((r['kg'] or 0) for r in detalle_por_plu))
        total_kg_dev = sum(r['kg_devuelto'] for r in detalle_enriquecido)
        totales_plus = {
            "detalle_retiros": detalle_enriquecido,
            "total_baldes_retirados": sum(r["cant"] for r in detalle_por_plu),
            "total_kg_retirados": total_kg_ret,
            "total_kg_devueltos": round(total_kg_dev, 3),
            "total_kg_neto": round(total_kg_ret - total_kg_dev, 3),
            "orden": orden,
        }

    # ------------------------------------------------------------
    # Resumen de INGRESOS por PLU
    # ------------------------------------------------------------
    qs_ingresos = base_qs.filter(tipo="ingreso")
    if plus_list:
        qs_ingresos = qs_ingresos.filter(producto__plu__in=plus_list)

    detalle_por_plu_ing_qs = (
        qs_ingresos
        .values("producto__plu", "producto__nombre")
        .annotate(cant=Count("id"), kg=Sum("peso"))
    )
    if orden_ing == "kg_asc":
        detalle_por_plu_ing_qs = detalle_por_plu_ing_qs.order_by("kg", "producto__nombre", "producto__plu")
    else:
        detalle_por_plu_ing_qs = detalle_por_plu_ing_qs.order_by("-kg", "producto__nombre", "producto__plu")

    detalle_por_plu_ing = list(detalle_por_plu_ing_qs)
    totales_plus_ing = None
    if detalle_por_plu_ing:
        # ✅ DESPUÉS (cambiar el nombre de la clave)
        totales_plus_ing = {
            "detalle_ingresos": detalle_por_plu_ing,  # ← Nombre específico para ingresos
            "total_baldes_ingresados": sum(r["cant"] for r in detalle_por_plu_ing),
            "total_kg_ingresados": sum((r["kg"] or 0) for r in detalle_por_plu_ing),
            "orden_ing": orden_ing,
        }

    # ----------------------------------------
    # Exportaciones (antes del render)
    # ----------------------------------------
    if export in {"xlsx", "ing_xlsx"}:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import io

        if export == "xlsx":
            sheet_title = "Retiros por PLU"
            rows = detalle_por_plu
            fname = "resumen_retiros_por_plu.xlsx"
        else:
            sheet_title = "Ingresos por PLU"
            rows = detalle_por_plu_ing
            fname = "resumen_ingresos_por_plu.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        headers = ["PLU", "Producto", "Cant. baldes", "Kilos"]
        ws.append(headers)

        for r in rows:
            ws.append([
                r.get("producto__plu", ""),
                r.get("producto__nombre", ""),
                r.get("cant", 0),
                float(r.get("kg") or 0),
            ])

        if rows:
            ws.append([])
            ws.append([
                "", "TOTAL",
                sum(x["cant"] for x in rows),
                float(sum((x["kg"] or 0) for x in rows)),
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    # -------- Render HTML --------
    return render(
        request,
        "historial_movimientos.html",
        {
            "modo": "historial",
            "movimientos": movimientos_page,
            "filtros": {
                "desde": desde_str,
                "hasta": hasta_str,
                "local": local,
                "tipo": tipo,
                "gusto": gusto,
                "codigo": codigo,
                "plus": plus_str,
                "orden": orden,
                "orden_ing": orden_ing,
                "solo_activos": False,
            },
            "total_kg_global": total_kg_global,
            "totales_plus": totales_plus,
            "totales_plus_ing": totales_plus_ing,
        },
    )


    


def detalle_movimiento(request, grupo_id: int):
    items = (RegistroMovimiento.objects
             .filter(grupo_id=grupo_id)
             .select_related("producto", "destino")
             .order_by("timestamp", "id"))

    if not items.exists():
        raise Http404("Movimiento no encontrado")

    header = GrupoMovimiento.objects.filter(grupo_id=grupo_id).select_related("destino").first()
    if header:
        total_peso = header.total_peso or 0
        cantidad_items = header.cantidad_items or items.count()
        tipo = header.tipo
        origen = header.origen
        destino = header.destino
    else:
        agg = items.aggregate(total=Sum("peso"), cantidad=Count("id"),
                              tipo_any=Max("tipo"),
                              origen_any=Max("origen"),
                              destino_any=Max("destino"))
        total_peso = agg["total"] or 0
        cantidad_items = agg["cantidad"] or 0
        tipo = agg["tipo_any"]
        origen = agg["origen_any"]
        destino = None
        if agg["destino_any"]:
            destino = BocaSalida.objects.filter(pk=agg["destino_any"]).first()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.GET.get("format") == "json":
        return JsonResponse({
            "grupo_id": grupo_id,
            "tipo": tipo,
            "origen": origen,
            "destino": destino.nombre if destino else None,
            "total_peso": round(float(total_peso), 2),
            "cantidad_items": int(cantidad_items),
            "items": [
                {
                    "id": i.id,
                    "producto": i.producto.nombre,
                    "plu": i.producto.plu,
                    "peso": float(i.peso),
                    "codigo_barras": getattr(i, "codigo_barras", None),
                }
                for i in items
            ],
        })

    return render(request, "detalle_movimiento.html", {
        "grupo_id": grupo_id,
        "tipo": tipo,
        "origen": origen,
        "destino": destino,
        "total_peso": total_peso,
        "cantidad_items": cantidad_items,
        "items": items,
    })


@csrf_exempt
def eliminar_movimiento(request, grupo_id):
    if request.method != "DELETE":
        return JsonResponse({"success": False, "error": "Método no permitido."}, status=405)

    # Cargar registros con el balde asociado en una sola query
    movs = RegistroMovimiento.objects.filter(grupo_id=grupo_id).select_related("balde")
    if not movs.exists():
        return JsonResponse({"success": False, "error": "Movimiento no encontrado."}, status=404)

    tipo = movs.values_list("tipo", flat=True).first()

    # Fix BUG-B: excepción local para salir del atomic con ROLLBACK garantizado.
    # `raise` fuerza el rollback antes de que el except exterior devuelva el error.
    # `return` dentro de `with transaction.atomic():` haría COMMIT de las
    # escrituras parciales (balde.delete()) ejecutadas en iteraciones previas.
    class _ErrorAnulacion(Exception):
        def __init__(self, mensaje, status_code=400):
            self.mensaje = mensaje
            self.status_code = status_code

    try:
        with transaction.atomic():
            if tipo in ("ingreso", "devolucion"):
                # Para ingreso y devolución: eliminar el StockBalde creado por este movimiento.
                # Si el balde_id está disponible → acceso directo y exacto al balde.
                # Si no (registro histórico) → fallback FIFO por codigo_barras.
                for mov in movs:
                    if mov.balde_id:
                        # FK directo: el balde exacto involucrado en este movimiento
                        balde = mov.balde
                        if not balde:
                            # Balde ya no existe (eliminado externamente) — skip
                            continue
                        if not balde.is_activo:
                            raise _ErrorAnulacion(
                                "No se puede anular: uno o más baldes de este movimiento ya fueron retirados."
                            )
                        balde.delete()
                    elif mov.codigo_barras:
                        # Fallback para registros sin balde_id (datos históricos)
                        balde = (
                            StockBalde.objects
                            .filter(codigo_barras=mov.codigo_barras, is_activo=True)
                            .order_by("timestamp", "id")
                            .first()
                        )
                        if not balde:
                            raise _ErrorAnulacion(
                                "No se puede anular: uno o más baldes ya fueron retirados."
                            )
                        balde.delete()

            elif tipo == "salida":
                # Para salida (retiro): reactivar el StockBalde que fue retirado.
                # balde_id → reactivar el balde exacto.
                # Sin balde_id → reactivar el más recientemente retirado con ese código (fallback).
                for mov in movs:
                    if mov.balde_id:
                        balde = mov.balde
                        if balde:
                            balde.is_activo = True
                            balde.fecha_retiro = None
                            balde.save(update_fields=["is_activo", "fecha_retiro"])
                    elif mov.codigo_barras:
                        balde = (
                            StockBalde.objects
                            .filter(codigo_barras=mov.codigo_barras, is_activo=False)
                            .order_by("-fecha_retiro", "-id")
                            .first()
                        )
                        if balde:
                            balde.is_activo = True
                            balde.fecha_retiro = None
                            balde.save(update_fields=["is_activo", "fecha_retiro"])

            movs.delete()
            GrupoMovimiento.objects.filter(grupo_id=grupo_id).delete()

    except _ErrorAnulacion as exc:
        return JsonResponse({"success": False, "error": exc.mensaje}, status=exc.status_code)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error al eliminar: {e}"}, status=500)

    return JsonResponse({"success": True, "message": "Movimiento anulado correctamente."})


@csrf_exempt
def eliminar_item_movimiento(request):
    """DELETE /api/eliminar_item_movimiento/  { "registro_id": 123 }
    Elimina un balde individual dentro de un grupo de movimiento y ajusta los totales.
    """
    if request.method != "DELETE":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    registro_id = data.get("registro_id")
    if not registro_id:
        return JsonResponse({"error": "registro_id requerido"}, status=400)

    try:
        mov = RegistroMovimiento.objects.select_related("balde").get(id=registro_id)
    except RegistroMovimiento.DoesNotExist:
        return JsonResponse({"error": "Registro no encontrado"}, status=404)

    grupo_id = mov.grupo_id
    tipo = mov.tipo
    codigo = (mov.codigo_barras or "").strip()

    try:
        with transaction.atomic():
            if tipo in ("ingreso", "devolucion"):
                if mov.balde_id:
                    # FK directo: balde exacto involucrado en este ítem
                    balde = mov.balde
                    if not balde:
                        pass  # ya fue eliminado externamente — OK
                    elif not balde.is_activo:
                        return JsonResponse({
                            "error": "Este balde ya fue retirado y no puede anularse."
                        }, status=400)
                    else:
                        balde.delete()
                elif codigo:
                    # Fallback para registros históricos sin balde_id
                    if not StockBalde.objects.filter(codigo_barras=codigo, is_activo=True).exists():
                        return JsonResponse({
                            "error": "Este balde ya fue retirado y no puede anularse."
                        }, status=400)
                    balde = (
                        StockBalde.objects
                        .filter(codigo_barras=codigo, is_activo=True)
                        .order_by("timestamp", "id")
                        .first()
                    )
                    if balde:
                        balde.delete()

            elif tipo == "salida":
                if mov.balde_id:
                    # FK directo: reactivar el balde exacto que fue retirado
                    balde = mov.balde
                    if balde:
                        balde.is_activo = True
                        balde.fecha_retiro = None
                        balde.save(update_fields=["is_activo", "fecha_retiro"])
                elif codigo:
                    # Fallback: reactivar el más recientemente retirado con ese barcode
                    balde = (
                        StockBalde.objects
                        .filter(codigo_barras=codigo, is_activo=False)
                        .order_by("-fecha_retiro", "-id")
                        .first()
                    )
                    if balde:
                        balde.is_activo = True
                        balde.fecha_retiro = None
                        balde.save(update_fields=["is_activo", "fecha_retiro"])

            mov.delete()

            restantes = RegistroMovimiento.objects.filter(grupo_id=grupo_id)
            if restantes.exists():
                agg = restantes.aggregate(total=Sum("peso"), cant=Count("id"))
                GrupoMovimiento.objects.filter(grupo_id=grupo_id).update(
                    total_peso=agg["total"] or 0,
                    cantidad_items=agg["cant"] or 0,
                )
                items_restantes = agg["cant"] or 0
            else:
                GrupoMovimiento.objects.filter(grupo_id=grupo_id).delete()
                items_restantes = 0

    except Exception as e:
        return JsonResponse({"error": f"Error al eliminar: {e}"}, status=500)

    return JsonResponse({
        "success": True,
        "items_restantes": items_restantes,
        "message": "Balde eliminado del movimiento.",
    })


@csrf_exempt
def api_editar_item_movimiento(request):
    """POST /api/editar_item_movimiento/  { "registro_id": 123, "nuevo_plu": "001", "nuevo_peso": 2.500 }
    Corrige el producto y/o peso de un balde dentro de un grupo de movimiento.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    registro_id = data.get("registro_id")
    nuevo_plu = (str(data.get("nuevo_plu") or "")).strip().zfill(3)
    nuevo_peso_raw = data.get("nuevo_peso")

    if not registro_id:
        return JsonResponse({"error": "registro_id requerido"}, status=400)

    try:
        nuevo_peso = float(nuevo_peso_raw)
        if nuevo_peso <= 0:
            return JsonResponse({"error": "El peso debe ser mayor a 0"}, status=400)
    except (TypeError, ValueError):
        return JsonResponse({"error": "Peso inválido"}, status=400)

    try:
        producto = ProductoFijo.objects.get(plu=nuevo_plu)
    except ProductoFijo.DoesNotExist:
        return JsonResponse({"error": f"No existe producto con PLU {nuevo_plu}"}, status=404)

    try:
        registro = RegistroMovimiento.objects.select_related("destino", "balde").get(id=registro_id)
    except RegistroMovimiento.DoesNotExist:
        return JsonResponse({"error": "Registro no encontrado"}, status=404)

    with transaction.atomic():
        if registro.tipo in ("ingreso", "devolucion"):
            if registro.balde_id:
                # FK directo: actualizar el balde exacto de este movimiento
                balde = registro.balde
                if balde and balde.is_activo:
                    balde.producto = producto
                    balde.peso = nuevo_peso
                    balde.save(update_fields=["producto", "peso"])
            elif registro.codigo_barras:
                # Fix Bug #8: actualizar SOLO el balde más antiguo activo (FIFO),
                # no todos los que compartan el mismo barcode.
                balde_fallback = (
                    StockBalde.objects
                    .filter(codigo_barras=registro.codigo_barras, is_activo=True)
                    .order_by("timestamp", "id")
                    .first()
                )
                if balde_fallback:
                    balde_fallback.producto = producto
                    balde_fallback.peso = nuevo_peso
                    balde_fallback.save(update_fields=["producto", "peso"])
        # Para salidas: el balde ya está INACTIVO (retirado). No tocar StockBalde,
        # solo corregir el RegistroMovimiento para trazabilidad del movimiento.

        registro.producto = producto
        registro.peso = nuevo_peso
        registro.save(update_fields=["producto", "peso"])

        destino_nombre = registro.destino.nombre if registro.destino else None
        _actualizar_total_grupo(
            registro.grupo_id,
            tipo=registro.tipo,
            origen=registro.origen,
            destino_nombre=destino_nombre,
        )

    return JsonResponse({"success": True, "message": "Balde corregido correctamente"})


def buscar(request):
    return render(request, "index.html")


def buscar_detallado(request):
    if request.method == "POST":
        termino_busqueda = request.POST.get("termino_busqueda", "").strip()
        fecha = request.POST.get("fecha", "").strip()

        if not termino_busqueda and not fecha:
            productos = StockBalde.objects.all()
        else:
            productos = StockBalde.objects.filter(
                Q(producto__nombre__icontains=termino_busqueda) |
                Q(producto__plu__icontains=termino_busqueda)
            )
            fecha_formateada = parse_date(fecha)
            if fecha_formateada:
                productos = productos.filter(timestamp__date=fecha_formateada)

        return render(request, "stock_detallado.html", {"stock_detallado": productos})

    return JsonResponse({"error": "Método no permitido"}, status=405)


# =========================================================
# API: lista temporal por sesión (escaneo)
# =========================================================

@csrf_exempt
def reiniciar_lista_temporal(request):
    request.session["productos_temporales"] = []
    request.session.modified = True
    return JsonResponse({"message": "Lista de productos escaneados reiniciada"})


@csrf_exempt
def obtener_codigos(request):
    return JsonResponse({"productos": request.session.get("productos_temporales", [])})


@csrf_exempt
def procesar_codigo(request):
    """
    Recibe {"codigo": "xxxxxxxxxxxxx"} (EAN-13).
    - Extrae PLU (3 dígitos) y peso (X.XXX) del código.
    - Verifica que el producto exista.
    - Agrega a la lista temporal en sesión incluyendo `codigo_barras`.
    - Evita duplicados por el mismo `codigo_barras`.
    Responde con la lista temporal actualizada.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)

    # --- Parseo de body ---
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Formato JSON inválido"}, status=400)

    codigo_barras = (data.get("codigo") or "").strip()

    # Validación básica de EAN-13 (longitud y dígitos)
    if not (len(codigo_barras) == 13 and codigo_barras.isdigit()):
        return JsonResponse({"error": "Código de barras no válido"}, status=400)

    # --- Interpretación del código ---
    plu = codigo_barras[2:5]  # 3 dígitos
    peso = float(f"{codigo_barras[8]}.{codigo_barras[9:12]}")  # X.XXX

    # --- Producto existente ---
    try:
        prod = ProductoFijo.objects.get(plu=plu)
    except ProductoFijo.DoesNotExist:
        return JsonResponse({"error": "Producto no encontrado"}, status=400)

    # --- Lista temporal en sesión ---
    productos_temporales = request.session.get("productos_temporales", [])

    # Armamos el item incluyendo el código de barras
    nuevo = {
        "nombre": prod.nombre,
        "peso": peso,
        "plu": plu,
        "codigo_barras": codigo_barras,  # 👈 clave nueva para trazabilidad
    }

    # Evitar duplicados: si ya existe ese mismo código en la lista, no lo agregamos
    if not any(p.get("codigo_barras") == codigo_barras for p in productos_temporales):
        productos_temporales.append(nuevo)

    # Persistir en sesión
    request.session["productos_temporales"] = productos_temporales
    request.session.modified = True

    return JsonResponse({
        "message": "Producto agregado temporalmente",
        "productos_temporales": productos_temporales
    })


@csrf_exempt
def obtener_productos_temporales(request):
    return JsonResponse({"productos": request.session.get("productos_temporales", [])})


@csrf_exempt
def eliminar_producto_temporal(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
        codigo_barras = (data.get("codigo_barras") or "").strip()
        if not codigo_barras:
            return JsonResponse({"success": False, "error": "codigo_barras requerido"}, status=400)
        productos_temporales = request.session.get("productos_temporales", [])
        # Eliminar SOLO el primer elemento con ese codigo_barras exacto (no todos los del mismo PLU)
        nuevo_listado = []
        eliminado = False
        for p in productos_temporales:
            if not eliminado and p.get("codigo_barras") == codigo_barras:
                eliminado = True  # saltar solo la primera ocurrencia
            else:
                nuevo_listado.append(p)
        request.session["productos_temporales"] = nuevo_listado
        request.session.modified = True
        return JsonResponse({"success": True, "message": "Producto eliminado de la sesión."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
def confirmar_agregado(request):
    """
    Vuelca la lista temporal al stock (sin crear grupo/movimientos).
    Mantenida por compatibilidad: si preferís trabajar 100% con grupos, podés prescindir de esta.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)

    productos = request.session.get("productos_temporales", [])
    if not productos:
        return JsonResponse({"error": "No hay productos para confirmar"}, status=400)

    productos_agregados = []
    for p in productos:
        try:
            producto_obj = ProductoFijo.objects.get(plu=p["plu"])
            StockBalde.objects.create(producto=producto_obj, peso=p["peso"])
            productos_agregados.append(producto_obj.nombre)
        except ProductoFijo.DoesNotExist:
            return JsonResponse({"error": f"Producto con PLU {p['plu']} no encontrado"}, status=404)

    request.session["productos_temporales"] = []
    request.session.modified = True

    return JsonResponse({"message": f"Productos agregados: {', '.join(productos_agregados)} ✅"}, status=200)


# =========================================================
# Confirmar Ingreso / Retiro (agrupado con grupo_id)
# =========================================================
@csrf_exempt
def confirmar_codigos(request):
    """
    Confirma un INGRESO de baldes con trazabilidad por código de barras.

    Mejoras anti-duplicados:
    - nuevo_grupo_id se calcula dentro de transaction y serializado
    - opcional: client_txn_id (idempotencia) guardado en sesión
    - chequeo de duplicado con select_for_update para evitar carreras
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    # --- Parseo payload ---
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Formato JSON inválido"}, status=400)

    productos = data.get("productos", []) or request.session.get("productos_temporales", [])
    origen = (data.get("origen") or "").strip()
    force = bool(data.get("force", False))

    # ✅ Token opcional para hacer la operación idempotente (recomendado)
    # En el front mandás un UUID por cada confirmación (una sola vez).
    client_txn_id = (data.get("client_txn_id") or "").strip()

    if not productos:
        return JsonResponse({"error": "No hay productos para ingresar"}, status=400)
    if not origen:
        return JsonResponse({"error": "Debe indicar un origen"}, status=400)

    # ---- Idempotencia por sesión (opcional) ----
    # Evita que si el usuario confirma 2 veces, la segunda repita todo.
    if client_txn_id:
        processed = request.session.get("processed_txn_ids", [])
        if client_txn_id in processed:
            # Ya procesado: devolvemos OK sin volver a ingresar.
            return JsonResponse({
                "success": True,
                "status": "ya_procesado",
                "message": "Esta confirmación ya fue procesada.",
            }, status=200)

    ingresados = []

    # ── Fix Bug #1: separar validación de escritura ────────────────────────────
    # Fase 1 (pre-validación, fuera de transacción): si algún producto es inválido
    # se rechaza todo ANTES de tocar la BD → no quedan baldes huérfanos.
    validados = []
    for p in productos:
        plu = (p or {}).get("plu")
        peso = (p or {}).get("peso")
        codigo_barras = (p or {}).get("codigo_barras") or (p or {}).get("codigo")

        if not plu:
            return JsonResponse({"error": "Producto sin PLU"}, status=400)
        if peso in (None, "", 0):
            return JsonResponse({"error": f"Producto {plu} sin peso"}, status=400)

        codigo_str = str(codigo_barras or "")
        if len(codigo_str) != 13 or not codigo_str.isdigit():
            return JsonResponse(
                {"error": "Cada balde debe incluir 'codigo_barras' de 13 dígitos"},
                status=400,
            )

        try:
            producto_obj = ProductoFijo.objects.get(plu=plu)
        except ProductoFijo.DoesNotExist:
            return JsonResponse({"error": f"Producto con PLU {plu} no encontrado"}, status=404)

        validados.append({
            "producto": producto_obj,
            "peso": float(peso),
            "codigo": codigo_str,
        })

    # Excepción local para salir limpiamente del bloque atómico al detectar duplicado.
    # `raise` dentro de `with transaction.atomic()` garantiza el rollback antes de que
    # el except exterior devuelva la respuesta 409.
    class _DuplicadoDetectado(Exception):
        def __init__(self, data, status_code=409):
            self.data = data
            self.status_code = status_code

    # Fase 2 (transacción): solo escrituras + chequeo de duplicados con lock
    try:
        with transaction.atomic():
            # ✅ Serializar la generación del grupo_id para evitar carreras
            # Bloquea la fila más "alta" de grupo_id momentáneamente.
            ultimo = (
                RegistroMovimiento.objects
                .select_for_update()
                .order_by("-grupo_id")
                .values_list("grupo_id", flat=True)
                .first()
            )
            ultimo_grupo = ultimo or 0
            nuevo_grupo_id = ultimo_grupo + 1

            for item in validados:
                producto_obj = item["producto"]
                peso_f = item["peso"]
                codigo_str = item["codigo"]

                # ✅ Anti-carrera: bloqueamos cualquier balde existente con ese código
                # Si otro request está tratando el mismo código, uno va a esperar al otro.
                duplicado_qs = (
                    StockBalde.objects
                    .select_for_update()
                    .filter(codigo_barras=codigo_str, is_activo=True)
                )

                if duplicado_qs.exists() and not force:
                    ultimo_dup = (
                        duplicado_qs.order_by("-id")
                        .values("producto__nombre", "peso", "fecha_retiro", "timestamp")
                        .first()
                    )
                    # raise → rollback automático del atomic block
                    raise _DuplicadoDetectado(
                        {
                            "status": "duplicado_detectado",
                            "codigo_barras": codigo_str,
                            "producto": ultimo_dup.get("producto__nombre") if ultimo_dup else None,
                            "peso_anterior": ultimo_dup.get("peso") if ultimo_dup else None,
                            "fecha_retiro": ultimo_dup.get("fecha_retiro") if ultimo_dup else None,
                            "fecha_ingreso": ultimo_dup.get("timestamp").isoformat()
                                if ultimo_dup and ultimo_dup.get("timestamp") else None,
                            "mensaje": f"⚠️ Este balde con código <b>{codigo_str}</b> "
                                       f"ya se encuentra <b>ACTIVO</b> en el stock.<br><br>",
                            "se_puede_forzar": True,
                        },
                        status_code=409,
                    )

                # ✅ Crear balde activo
                balde = StockBalde.objects.create(
                    producto=producto_obj,
                    peso=peso_f,
                    codigo_barras=codigo_str,
                    is_activo=True,
                    fecha_retiro=None,
                )

                # ✅ Registrar movimiento con referencia directa al balde (balde_id)
                RegistroMovimiento.objects.create(
                    grupo_id=nuevo_grupo_id,
                    producto=producto_obj,
                    peso=peso_f,
                    tipo="ingreso",
                    origen=origen,
                    boca_salida=origen,  # compatibilidad con layouts existentes
                    codigo_barras=codigo_str,
                    balde=balde,
                )

                ingresados.append(producto_obj.nombre)

            # ✅ Totales del grupo
            _actualizar_total_grupo(nuevo_grupo_id, tipo="ingreso", origen=origen)

            # ✅ Limpiar sesión temporal
            if "productos_temporales" in request.session:
                request.session["productos_temporales"] = []
                request.session.modified = True

            # ✅ Marcar txn_id como procesado (idempotencia)
            if client_txn_id:
                processed = request.session.get("processed_txn_ids", [])
                processed.append(client_txn_id)
                # evito crecimiento infinito
                request.session["processed_txn_ids"] = processed[-50:]
                request.session.modified = True

    except _DuplicadoDetectado as exc:
        return JsonResponse(exc.data, status=exc.status_code)
    except Exception as e:
        return JsonResponse({"error": f"Error al confirmar ingreso: {e}"}, status=500)

    msg = "Productos agregados correctamente:\n\n" + "\n".join(ingresados)

    return JsonResponse({
        "success": True,
        "grupo_id": nuevo_grupo_id,
        "origen": origen,
        "productos": ingresados,
        "message": msg
    }, status=200)


@csrf_exempt
def confirmar_retiro(request):
    """
    Confirma RETIRO de baldes.

    Escenario actual:
      - TODOS los baldes tienen código de barras.
      - Para cada ítem se espera: {plu, codigo_barras}
      - No hay baldes "legacy" sin código.
      - Si el mismo (plu, codigo_barras) viene repetido en el payload,
        sólo se procesa UNA vez (deduplicación en backend).
      - Si no existe un balde ACTIVO con ese código, se informa como faltante.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    # -------- Parseo body --------
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Formato JSON inválido"}, status=400)

    productos = data.get("productos", [])
    destino_nombre = (data.get("destino") or "").strip()

    if not productos:
        return JsonResponse({"error": "No hay productos para retirar"}, status=400)
    if not destino_nombre:
        return JsonResponse({"error": "Debe indicar un destino"}, status=400)

    # -------- Destino (FK) --------
    destino_obj = BocaSalida.objects.filter(nombre=destino_nombre).first()
    if not destino_obj:
        return JsonResponse({"error": f"Destino '{destino_nombre}' no existe"}, status=400)

    # -------- Normalizar solicitudes y EVITAR DUPLICADOS --------
    # Generamos una lista de (plu, codigo) única
    solicitudes = []
    vistos = set()   # set de (plu, codigo_barras)

    for p in productos:
        p = p or {}
        plu = p.get("plu")
        codigo = (p.get("codigo_barras") or p.get("codigo") or "").strip()

        if not plu:
            return JsonResponse({"error": "Producto sin PLU"}, status=400)
        if not codigo:
            return JsonResponse({"error": f"Producto {plu} sin código de barras"}, status=400)
        if len(codigo) != 13 or not codigo.isdigit():
            return JsonResponse({"error": f"Código inválido para PLU {plu}: debe ser EAN-13"}, status=400)

        clave = (plu, codigo)
        if clave in vistos:
            # Ya tenemos este mismo balde en la lista, lo ignoramos
            continue
        vistos.add(clave)
        solicitudes.append((plu, codigo))

    # -------- Fase 1: lookup de PLU y pre-verificación de stock (sin lock) --------
    # La re-verificación definitiva ocurre dentro del atomic con select_for_update (Fase 2).
    solicitudes_validadas = []  # [(producto_obj, codigo_barras), ...]
    faltantes = []

    for plu, codigo in solicitudes:
        try:
            producto_obj = ProductoFijo.objects.get(plu=plu)
        except ProductoFijo.DoesNotExist:
            return JsonResponse({"error": f"Producto con PLU {plu} no encontrado"}, status=404)

        # Pre-check rápido (sin select_for_update) para el caso habitual (sin concurrencia)
        if StockBalde.objects.filter(
            producto=producto_obj, is_activo=True, codigo_barras=codigo
        ).exists():
            solicitudes_validadas.append((producto_obj, codigo))
        else:
            faltantes.append(f"{producto_obj.nombre} ({codigo})")

    if faltantes:
        return JsonResponse(
            {"error": f"No hay stock disponible para: {', '.join(faltantes)}"},
            status=400,
        )

    # -------- Fase 2: retiro atómico con locks (corrige BUG-1 y BUG-2) --------
    # BUG-1: el balde se re-lee con select_for_update dentro del atomic →
    #        si otro proceso lo retiró entre el pre-check y aquí, se detecta y rollback.
    # BUG-2: el grupo_id se calcula con select_for_update dentro del atomic →
    #        no puede colisionar con el de un request concurrente.

    class _ErrorRetiro(Exception):
        def __init__(self, data, status_code=400):
            self.data = data
            self.status_code = status_code

    productos_retirados = []
    nuevo_grupo_id = None
    try:
        with transaction.atomic():
            # grupo_id serializado dentro del atomic (mismo patrón que confirmar_codigos)
            ultimo = (
                RegistroMovimiento.objects
                .select_for_update()
                .order_by("-grupo_id")
                .values_list("grupo_id", flat=True)
                .first()
            )
            nuevo_grupo_id = (ultimo or 0) + 1

            for producto_obj, codigo in solicitudes_validadas:
                # Re-fetch con lock: detecta retiros concurrentes entre el pre-check y aquí
                balde = (
                    StockBalde.objects
                    .select_for_update()
                    .filter(producto=producto_obj, is_activo=True, codigo_barras=codigo)
                    .order_by("timestamp", "id")
                    .first()
                )
                if not balde:
                    raise _ErrorRetiro(
                        {"error": (
                            f"El balde {codigo} ({producto_obj.nombre}) "
                            f"fue retirado por otro proceso."
                        )},
                        status_code=409,
                    )

                # Marcar balde como inactivo
                balde.is_activo = False
                balde.fecha_retiro = timezone.now()
                balde.save(update_fields=["is_activo", "fecha_retiro"])

                # Registrar movimiento con referencia directa al balde (balde_id)
                RegistroMovimiento.objects.create(
                    grupo_id=nuevo_grupo_id,
                    producto=producto_obj,
                    peso=balde.peso,
                    tipo="salida",
                    destino=destino_obj,
                    boca_salida=destino_nombre,
                    codigo_barras=(balde.codigo_barras or ""),
                    balde=balde,
                )
                productos_retirados.append(producto_obj.nombre)

            # Totales del grupo
            _actualizar_total_grupo(
                nuevo_grupo_id,
                tipo="salida",
                destino_nombre=destino_nombre,
            )

    except _ErrorRetiro as exc:
        return JsonResponse(exc.data, status=exc.status_code)
    except Exception as e:
        return JsonResponse({"error": f"Error al retirar productos: {e}"}, status=500)

    msg = "Productos retirados correctamente:\n\n" + "\n".join(productos_retirados)
    return JsonResponse(
        {
            "success": True,
            "grupo_id": nuevo_grupo_id,
            "destino": destino_nombre,
            "productos": productos_retirados,
            "message": msg,
        },
        status=200,
    )

# =========================================================
# Devolución de baldes
# =========================================================

@csrf_exempt
def confirmar_devolucion(request):
    """
    POST /api/confirmar_devolucion/
    Body: { "productos": [...], "origen": "Local Norte" }

    El balde fue re-etiquetado: el nuevo código ya contiene producto y peso.
    Se crea un StockBalde nuevo (como en ingreso) pero con tipo='devolucion'.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Formato JSON inválido"}, status=400)

    productos = data.get("productos", []) or request.session.get("productos_temporales", [])
    origen  = (data.get("origen")  or "").strip()
    destino = (data.get("destino") or "").strip()   # boca destino para redirigir (opcional)

    if not productos:
        return JsonResponse({"error": "No hay baldes para devolver"}, status=400)

    # ── Fix BUG-A: separar validación de escritura ────────────────────────────
    # Excepción local para salir del bloque atómico con rollback garantizado.
    # `raise` dentro de `with transaction.atomic()` hace ROLLBACK antes de que
    # el except exterior devuelva la respuesta de error.
    class _ErrorDevolucion(Exception):
        def __init__(self, data, status_code=400):
            self.data = data
            self.status_code = status_code

    # Fase 1 (pre-validación, FUERA de la transacción):
    # Si algún item es inválido se rechaza TODO antes de tocar la BD.
    validados_dev = []
    codigos_devolucion_vistos = set()
    for p in productos:
        plu        = (p or {}).get("plu")
        peso       = (p or {}).get("peso")
        codigo_str = str((p or {}).get("codigo_barras") or "")

        if not plu or peso in (None, "", 0):
            return JsonResponse({"error": "Datos incompletos en la lista"}, status=400)
        if len(codigo_str) != 13 or not codigo_str.isdigit():
            return JsonResponse({"error": "Código de barras inválido"}, status=400)

        if codigo_str in codigos_devolucion_vistos:
            continue
        codigos_devolucion_vistos.add(codigo_str)

        try:
            producto_obj = ProductoFijo.objects.get(plu=plu)
        except ProductoFijo.DoesNotExist:
            return JsonResponse({"error": f"Producto PLU {plu} no encontrado"}, status=404)

        validados_dev.append({
            "producto": producto_obj,
            "peso": float(peso),
            "codigo": codigo_str,
        })

    # Fase 2 (transacción): solo escrituras + chequeo de doble-devolución con lock.
    devueltos = []
    grupo_id_retiro = None
    try:
        with transaction.atomic():
            ultimo = (
                RegistroMovimiento.objects
                .select_for_update()
                .order_by("-grupo_id")
                .values_list("grupo_id", flat=True)
                .first()
            )
            nuevo_grupo_id = (ultimo or 0) + 1

            baldes_creados = []   # para el retiro encadenado si hay destino

            for item in validados_dev:
                producto_obj = item["producto"]
                peso_f       = item["peso"]
                codigo_str   = item["codigo"]

                # Doble-devolución: debe verificarse dentro del atomic para evitar
                # carreras entre requests concurrentes.
                if StockBalde.objects.filter(codigo_barras=codigo_str, is_activo=True).exists():
                    raise _ErrorDevolucion(
                        {"error": f"El balde con código {codigo_str} ya está en stock activo."},
                        status_code=409,
                    )

                balde_dev = StockBalde.objects.create(
                    producto=producto_obj,
                    peso=peso_f,
                    codigo_barras=codigo_str,
                    is_activo=True,
                )
                RegistroMovimiento.objects.create(
                    grupo_id=nuevo_grupo_id,
                    producto=producto_obj,
                    peso=peso_f,
                    tipo="devolucion",
                    origen=origen or None,
                    codigo_barras=codigo_str,
                    balde=balde_dev,
                )
                devueltos.append(producto_obj.nombre)
                baldes_creados.append(balde_dev)

            _actualizar_total_grupo(nuevo_grupo_id, tipo="devolucion", origen=origen)

            # ── Redirigir: si hay destino, crear retiro encadenado ────────────
            if destino and baldes_creados:
                grupo_id_retiro = nuevo_grupo_id + 1
                ahora = timezone.now()
                destino_obj_retiro = BocaSalida.objects.filter(nombre=destino).first()
                for balde_dev in baldes_creados:
                    balde_dev.is_activo    = False
                    balde_dev.fecha_retiro = ahora
                    balde_dev.save(update_fields=["is_activo", "fecha_retiro"])
                    RegistroMovimiento.objects.create(
                        grupo_id    = grupo_id_retiro,
                        producto    = balde_dev.producto,
                        peso        = balde_dev.peso,
                        tipo        = "retiro",
                        destino     = destino_obj_retiro,   # FK — usado por el historial
                        boca_salida = destino,              # string — usado por filtros/reportes
                        codigo_barras = balde_dev.codigo_barras,
                        balde       = balde_dev,
                    )
                _actualizar_total_grupo(grupo_id_retiro, tipo="retiro", destino_nombre=destino)

            request.session["productos_temporales"] = []
            request.session.modified = True

    except _ErrorDevolucion as exc:
        return JsonResponse(exc.data, status=exc.status_code)
    except Exception as e:
        return JsonResponse({"error": f"Error al procesar devolución: {e}"}, status=500)

    if destino and grupo_id_retiro:
        msg = (
            f"{len(devueltos)} balde(s) devuelto(s) desde {origen} "
            f"y redirigido(s) a {destino}."
        )
    else:
        msg = f"Devolución registrada: {len(devueltos)} balde(s) reingresado(s)."

    return JsonResponse({
        "success":        True,
        "grupo_id":       nuevo_grupo_id,
        "grupo_id_retiro": grupo_id_retiro,
        "cantidad":       len(devueltos),
        "message":        msg,
    })


# =========================================================
# Catálogos: Bocas / Orígenes
# =========================================================

@csrf_exempt
def obtener_bocas(request):
    if request.method == "GET":
        bocas = list(BocaSalida.objects.values_list("nombre", flat=True))
        return JsonResponse({"bocas": bocas})
    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
def crear_boca(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return JsonResponse({"success": False, "error": "El nombre no puede estar vacío"}, status=400)
            if BocaSalida.objects.filter(nombre__iexact=nombre).exists():
                return JsonResponse({"success": False, "error": "Ya existe una boca con ese nombre"}, status=400)
            BocaSalida.objects.create(nombre=nombre)
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


@csrf_exempt
def crear_origen(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return JsonResponse({"success": False, "error": "El nombre no puede estar vacío"}, status=400)
            if OrigenIngreso.objects.filter(nombre__iexact=nombre).exists():
                return JsonResponse({"success": False, "error": "Ya existe un origen con ese nombre"}, status=400)
            OrigenIngreso.objects.create(nombre=nombre)
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


def obtener_bocas_salida(request):
    if request.method == "GET":
        bocas = list(BocaSalida.objects.values_list("nombre", flat=True).order_by("nombre"))
        return JsonResponse({"lista": bocas})


def obtener_origenes(request):
    if request.method == "GET":
        origenes = list(OrigenIngreso.objects.values_list("nombre", flat=True).order_by("nombre"))
        return JsonResponse({"lista": origenes})


@csrf_exempt
def eliminar_boca_salida(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return JsonResponse({"success": False, "error": "Nombre no proporcionado"}, status=400)
            boca = BocaSalida.objects.filter(nombre__iexact=nombre).first()
            if boca:
                boca.delete()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Boca no encontrada"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


@csrf_exempt
def eliminar_origen(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return JsonResponse({"success": False, "error": "Nombre no proporcionado"}, status=400)
            origen = OrigenIngreso.objects.filter(nombre__iexact=nombre).first()
            if origen:
                origen.delete()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "error": "Origen no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


# =========================================================
# Backups y mantenimiento
# =========================================================

# views.py
import tempfile
import shutil
from django.db import connection, connections

def descargar_backup(request):
    """
    Descarga la base actual como backup.sqlite3 (con nombre con fecha).
    """
    db_path = settings.DATABASES["default"]["NAME"]
    if not os.path.exists(db_path):
        return JsonResponse({"error": "No se encontró la base de datos"}, status=404)

    # nombre con timestamp (opcional)
    from datetime import datetime
    fname = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sqlite3"
    return FileResponse(open(db_path, 'rb'), as_attachment=True, filename=fname)



# views.py
@csrf_exempt
def importar_backup(request):
    """
    Sube un backup .sqlite3 y reemplaza la DB de forma segura:
    - guarda a un archivo temporal
    - cierra conexiones
    - reemplaza atómicamente
    - corre migraciones (sin run_syncdb)
    """
    if request.method != "POST" or "archivo" not in request.FILES:
        return JsonResponse({"error": "❌ Método no permitido o archivo no enviado"}, status=400)

    up = request.FILES["archivo"]

    # Validaciones básicas
    if not up.name.lower().endswith(".sqlite3"):
        return JsonResponse({"success": False, "error": "El archivo debe ser .sqlite3"}, status=400)
    if up.size and up.size > 50 * 1024 * 1024:  # 50 MB
        return JsonResponse({"success": False, "error": "El archivo es demasiado grande"}, status=400)

    db_path = settings.DATABASES["default"]["NAME"]
    db_dir = os.path.dirname(db_path) or "."

    tmp_path = None
    try:
        # 1) Escribir el archivo subido a un temporal en el mismo directorio
        os.makedirs(db_dir, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(prefix="upload_", suffix=".sqlite3", dir=db_dir)
        os.close(fd)

        with open(tmp_path, "wb") as destino:
            for chunk in up.chunks():
                destino.write(chunk)

        # 2) Cerrar TODAS las conexiones (incluye APScheduler y otros hilos)
        connections.close_all()

        # 3) Backup previo opcional (por si querés volver atrás)
        prev_backup = None
        if os.path.exists(db_path):
            prev_backup = db_path + ".prev"
            try:
                if os.path.exists(prev_backup):
                    os.remove(prev_backup)
            except Exception:
                pass
            try:
                shutil.copy2(db_path, prev_backup)
            except Exception:
                prev_backup = None

        # 4) Reemplazar la base de datos
        # En Windows, os.replace falla si el archivo destino sigue bloqueado.
        # Usamos shutil.copy2 + borrado del original como fallback seguro.
        try:
            os.replace(tmp_path, db_path)
        except OSError:
            # Fallback: copiar encima y eliminar el temporal
            shutil.copy2(tmp_path, db_path)
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            tmp_path = None

        # 5) Migraciones (sin run_syncdb)
        try:
            # Asegura que Django reabra la conexión contra la DB nueva
            connection.close()
            call_command("migrate", interactive=False, verbosity=0)
        except Exception as mig_e:
            # No abortamos: normalmente no hace falta si el backup ya estaba migrado
            print(f"[importar_backup] Warning al migrar: {mig_e}")

        # 6) Parche mínimo por SQL si faltara is_activo (defensivo)
        try:
            with connection.cursor() as c:
                c.execute("PRAGMA table_info(app_inventario_stockbalde);")
                cols = [r[1] for r in c.fetchall()]
                if "is_activo" not in cols:
                    c.execute(
                        "ALTER TABLE app_inventario_stockbalde "
                        "ADD COLUMN is_activo INTEGER NOT NULL DEFAULT 1;"
                    )
                    c.execute(
                        "CREATE INDEX IF NOT EXISTS stockbalde_is_activo_idx "
                        "ON app_inventario_stockbalde(is_activo);"
                    )
        except Exception as patch_e:
            print(f"[importar_backup] Warning al parchear is_activo: {patch_e}")

        return JsonResponse({"success": True, "message": "Backup restaurado correctamente"})

    except Exception as e:
        # Limpieza si algo falla
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        return JsonResponse({"success": False, "error": str(e)}, status=500)





@csrf_exempt
def reiniciar_stock(request):
    if request.method == "POST":
        try:
            StockBalde.objects.all().delete()
            return JsonResponse({"success": True, "message": "✅ Todos los baldes fueron eliminados. El stock ahora está en cero."})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=500)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

# views.py
def dashboard(request):
    hoy = timezone.now().date()
    
    # Movimientos del día
    movimientos_hoy = RegistroMovimiento.objects.filter(
        timestamp__date=hoy
    ).aggregate(
        ingresos=Count('id', filter=Q(tipo='ingreso')),
        retiros=Count('id', filter=Q(tipo='salida'))
    )
    
    # Top 5 productos más rotados (últimos 30 días)
    hace_30_dias = timezone.now() - timedelta(days=30)
    top_productos = (RegistroMovimiento.objects
        .filter(timestamp__gte=hace_30_dias)
        .values('producto__nombre')
        .annotate(movimientos=Count('id'))
        .order_by('-movimientos')[:5]
    )
    
    # Productos bajo stock mínimo
    productos_bajo_stock = ProductoFijo.objects.annotate(
        stock_actual=Count('stockbalde', filter=Q(stockbalde__is_activo=True))
    ).filter(stock_actual__lt=F('stock_minimo'))
    
    context = {
        'movimientos_hoy': movimientos_hoy,
        'top_productos': top_productos,
        'productos_bajo_stock': productos_bajo_stock,
    }
    return render(request, 'dashboard.html', context)


def api_dashboard_metricas(request):
    """
    API que devuelve todas las métricas para el dashboard
    """
    
    try:
        # Obtener parámetros de fecha (opcional)
        fecha_desde_str = request.GET.get('desde')
        fecha_hasta_str = request.GET.get('hasta')
        hoy = timezone.now()
        inicio_dia = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
        hace_7_dias = hoy - timedelta(days=7)
        hace_30_dias = hoy - timedelta(days=30)
        inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # Filtro por categoría
        categoria = request.GET.get('categoria', None)
        _CAT_RANGES = {'helado': (1, 88), 'barra_torta': (89, 98), 'gastronomico': (100, 199)}
        if categoria not in _CAT_RANGES:
            categoria = None

        balde_plu_q = Q()
        mov_plu_q   = Q()
        prod_plu_q  = Q()
        if categoria:
            plu_lo, plu_hi = _CAT_RANGES[categoria]
            plu_ids = list(
                ProductoFijo.objects
                .annotate(plu_int=Cast('plu', IntegerField()))
                .filter(plu_int__gte=plu_lo, plu_int__lte=plu_hi)
                .values_list('plu', flat=True)
            )
            balde_plu_q = Q(producto__plu__in=plu_ids)
            mov_plu_q   = Q(producto__plu__in=plu_ids)
            prod_plu_q  = Q(plu__in=plu_ids)
        
        
        # ============================================================
        # 1. RESUMEN GENERAL
        # ============================================================
        
        # Stock actual (filtrado por categoría si corresponde)
        total_baldes = StockBalde.objects.filter(balde_plu_q, is_activo=True).count()
        total_kilos = StockBalde.objects.filter(balde_plu_q, is_activo=True).aggregate(
            total=Sum('peso')
        )['total'] or 0

        # Valor del inventario (ejemplo: cada kg vale $100)
        valor_por_kg = float(get_config('valor_inventario_kg', '22500'))
        valor_inventario = float(total_kilos) * valor_por_kg

        # Productos únicos en stock
        productos_en_stock = StockBalde.objects.filter(
            balde_plu_q, is_activo=True
        ).values('producto').distinct().count()
        
        # ============================================================
        # 2. MOVIMIENTOS DEL DÍA
        # ============================================================
        
        movimientos_hoy = RegistroMovimiento.objects.filter(
            mov_plu_q, timestamp__gte=inicio_dia
        ).aggregate(
            ingresos=Count('id', filter=Q(tipo='ingreso')),
            retiros=Count('id', filter=Q(tipo='salida')),
            kg_ingresados=Sum('peso', filter=Q(tipo='ingreso')),
            kg_retirados=Sum('peso', filter=Q(tipo='salida'))
        )
        
        # ============================================================
        # 3. TENDENCIAS (Comparación con ayer y semana pasada)
        # ============================================================
        
        # Ayer
        inicio_ayer = inicio_dia - timedelta(days=1)
        fin_ayer = inicio_dia
        
        movimientos_ayer = RegistroMovimiento.objects.filter(
            mov_plu_q, timestamp__gte=inicio_ayer, timestamp__lt=fin_ayer
        ).aggregate(
            ingresos=Count('id', filter=Q(tipo='ingreso')),
            retiros=Count('id', filter=Q(tipo='salida'))
        )
        
        # Calcular porcentajes de cambio
        def calcular_cambio(actual, anterior):
            if anterior == 0:
                return 100 if actual > 0 else 0
            return round(((actual - anterior) / anterior) * 100, 1)
        
        cambio_ingresos = calcular_cambio(
            movimientos_hoy['ingresos'] or 0,
            movimientos_ayer['ingresos'] or 0
        )
        
        cambio_retiros = calcular_cambio(
            movimientos_hoy['retiros'] or 0,
            movimientos_ayer['retiros'] or 0
        )
        
        # ============================================================
        # 4. TOP PRODUCTOS MÁS MOVIDOS (Últimos 30 días)
        # ============================================================
        
        top_productos = (
            RegistroMovimiento.objects
            .filter(mov_plu_q, timestamp__gte=hace_30_dias)
            .values('producto__nombre', 'producto__plu')
            .annotate(
                total_movimientos=Count('id'),
                total_kg=Sum('peso'),
                ingresos=Count('id', filter=Q(tipo='ingreso')),
                retiros=Count('id', filter=Q(tipo='salida'))
            )
            .order_by('-total_movimientos')[:10]
        )
        
        # ============================================================
        # 5. PRODUCTOS BAJO STOCK MÍNIMO
        # ============================================================
        
        productos_bajo_stock = ProductoFijo.objects.filter(prod_plu_q).annotate(
            stock_actual=Count('stockbalde', filter=Q(stockbalde__is_activo=True))
        ).filter(
            stock_actual__lt=F('stock_minimo')
        ).values(
            'plu', 'nombre', 'stock_minimo', 'stock_actual'
        ).order_by('stock_actual')
        
        # Calcular déficit
        productos_bajo_stock_list = []
        for p in productos_bajo_stock:
            deficit = p['stock_minimo'] - p['stock_actual']
            productos_bajo_stock_list.append({
                **p,
                'deficit': deficit,
                'porcentaje_stock': round((p['stock_actual'] / p['stock_minimo'] * 100), 1) if p['stock_minimo'] > 0 else 0
            })
        
        # ============================================================
        # 6. PRODUCTOS SIN MOVIMIENTO (Últimos 30 días)
        # ============================================================
        
        productos_con_movimiento = RegistroMovimiento.objects.filter(
            mov_plu_q, timestamp__gte=hace_30_dias
        ).values_list('producto_id', flat=True).distinct()

        productos_sin_movimiento = ProductoFijo.objects.filter(prod_plu_q).exclude(
            plu__in=productos_con_movimiento
        ).annotate(
            stock_actual=Count('stockbalde', filter=Q(stockbalde__is_activo=True))
        ).filter(stock_actual__gt=0).values(
            'plu', 'nombre', 'stock_actual'
        )[:10]
        
        # ============================================================
        # 7. GRÁFICO DE MOVIMIENTOS (Últimos 7 días)
        # ============================================================
        
        movimientos_7_dias = []
        for i in range(7):
            dia = inicio_dia - timedelta(days=6-i)
            dia_siguiente = dia + timedelta(days=1)
            
            movs = RegistroMovimiento.objects.filter(
                mov_plu_q, timestamp__gte=dia, timestamp__lt=dia_siguiente
            ).aggregate(
                ingresos=Count('id', filter=Q(tipo='ingreso')),
                retiros=Count('id', filter=Q(tipo='salida')),
                kg_ingresados=Sum('peso', filter=Q(tipo='ingreso')),
                kg_retirados=Sum('peso', filter=Q(tipo='salida'))
            )

            movimientos_7_dias.append({
                'fecha': dia.strftime('%d/%m'),
                'dia_completo': dia.strftime('%Y-%m-%d'),
                'ingresos': movs['ingresos'] or 0,
                'retiros': movs['retiros'] or 0,
                'kg_ingresados': float(movs['kg_ingresados'] or 0),
                'kg_retirados': float(movs['kg_retirados'] or 0)
            })

        # ============================================================
        # 7B. GRÁFICO DE MOVIMIENTOS ÚLTIMOS 30 DÍAS
        # ============================================================

        movimientos_30_dias = []
        for i in range(30):
            dia = inicio_dia - timedelta(days=29-i)  # Empezar desde hace 29 días
            dia_siguiente = dia + timedelta(days=1)
            
            movs = RegistroMovimiento.objects.filter(
                mov_plu_q, timestamp__gte=dia, timestamp__lt=dia_siguiente
            ).aggregate(
                ingresos=Count('id', filter=Q(tipo='ingreso')),
                retiros=Count('id', filter=Q(tipo='salida')),
                kg_ingresados=Sum('peso', filter=Q(tipo='ingreso')),
                kg_retirados=Sum('peso', filter=Q(tipo='salida'))
            )

            movimientos_30_dias.append({
                'dia': i + 1,
                'fecha': dia.strftime('%d/%m'),
                'fecha_completa': dia.strftime('%Y-%m-%d'),
                'ingresos': movs['ingresos'] or 0,
                'retiros': movs['retiros'] or 0,
                'kg_ingresados': float(movs['kg_ingresados'] or 0),
                'kg_retirados': float(movs['kg_retirados'] or 0),
                'es_hoy': dia.date() == hoy.date()
            })

        # Resumen de los últimos 30 días
        resumen_30_dias = {
            'total_dias': 30,
            'total_ingresos': sum(m['ingresos'] for m in movimientos_30_dias),
            'total_retiros': sum(m['retiros'] for m in movimientos_30_dias),
            'total_kg_ingresados': sum(m['kg_ingresados'] for m in movimientos_30_dias),
            'total_kg_retirados': sum(m['kg_retirados'] for m in movimientos_30_dias),
            'promedio_ingresos_dia': round(sum(m['ingresos'] for m in movimientos_30_dias) / 30, 1),
            'promedio_retiros_dia': round(sum(m['retiros'] for m in movimientos_30_dias) / 30, 1),
            'periodo': 'Últimos 30 días'
        }
        

        # ✅ NUEVO: Movimientos período custom
        movimientos_custom = None
        resumen_custom = None
        
        if fecha_desde_str and fecha_hasta_str:
            try:
                from datetime import datetime
                
                # Parsear fechas
                fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
                fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d')
                
                # Hacer aware (con timezone)
                fecha_desde = timezone.make_aware(fecha_desde.replace(hour=0, minute=0, second=0))
                fecha_hasta = timezone.make_aware(fecha_hasta.replace(hour=23, minute=59, second=59))
                
                # Calcular días
                diff_days = (fecha_hasta.date() - fecha_desde.date()).days + 1
                
                # Generar datos día por día
                movimientos_custom = []
                for i in range(diff_days):
                    dia = fecha_desde + timedelta(days=i)
                    dia_siguiente = dia + timedelta(days=1)
                    
                    movs = RegistroMovimiento.objects.filter(
                        mov_plu_q, timestamp__gte=dia, timestamp__lt=dia_siguiente
                    ).aggregate(
                        ingresos=Count('id', filter=Q(tipo='ingreso')),
                        retiros=Count('id', filter=Q(tipo='salida')),
                        kg_ingresados=Sum('peso', filter=Q(tipo='ingreso')),
                        kg_retirados=Sum('peso', filter=Q(tipo='salida'))
                    )
                    
                    movimientos_custom.append({
                        'dia': i + 1,
                        'fecha': dia.strftime('%d/%m'),
                        'fecha_completa': dia.strftime('%Y-%m-%d'),
                        'ingresos': movs['ingresos'] or 0,
                        'retiros': movs['retiros'] or 0,
                        'kg_ingresados': float(movs['kg_ingresados'] or 0),
                        'kg_retirados': float(movs['kg_retirados'] or 0),
                        'es_hoy': dia.date() == hoy.date()
                    })
                
                # Resumen del período custom
                resumen_custom = {
                    'total_dias': diff_days,
                    'total_ingresos': sum(m['ingresos'] for m in movimientos_custom),
                    'total_retiros': sum(m['retiros'] for m in movimientos_custom),
                    'total_kg_ingresados': sum(m['kg_ingresados'] for m in movimientos_custom),
                    'total_kg_retirados': sum(m['kg_retirados'] for m in movimientos_custom),
                    'periodo': f'{fecha_desde_str} a {fecha_hasta_str}'
                }
                
            except Exception as e:
                print(f"Error procesando fechas custom: {e}")

        # ============================================================
        # 8. DISTRIBUCIÓN POR GRUPO DE PRODUCTOS
        # ============================================================
        
        grupos_productos = {
            'jarabe': [
                "LIMON", "FRUTILLA AL AGUA", "DURAZNO"
            ],
            'chocolates': [
                "CHOCOLATE", "CHOCOLAE BLOCK", "CH. CABSHA", "CHOCO DUBAI",
                "AMARGO", "CH. ALMENDRAS", "CH. PASAS RHUM", "CHOCOLAT PORTOFINO",
                "CHOCOLATE INTENSO", "CHOCOLAT DEBILIDAD", "CHOC. BLANCO",
                "ROCHER", "TOFFEE BLANCO"
            ],
            'dulces': [
                "DCE LECHE", "DCE. LECHE NUEZ", "DCE. GRANIZADO", "SUPER DCE LECHE",
                "DCE. VAUQUITA", "D. LECHE PORTOFINO", "DCE. LECHE COOKIES",
                "BASE DULCE LECHE", "CHOCOTORTA"
            ],
            'blanca': [
                "AMERICANA", "VAINILLA", "TRAMONTANA", "GRANIZADO",
                "MENTA GRANIZADA", "CREMA FLAN", "FLAN MIXTO",
                "FRUTOS DEL BOSQUE", "CREMA DEL CIELO", "PANNACOTA",
                "MASCARPONE", "CHEESE CAKE", "CAPUCCINO", "OREO", "SNIKERS"
            ],
            'neutra': [
                "CEREZA", "PISTACHO", "FRUTILLA CREMA", "BANANA SPLIT",
                "MARACUYA", "ANANA AL CHANTILLY", "FRAMBUESA C/ CHOCO",
                "KINOTOS AL WHISKY", "DURAZNOS AL OPORTO", "MANZANA VERDE",
                "LEMON PIE", "LIMON C/MARACUYA", "FRAMBUESA C/CHOCO",
                "HAVANETA LIMON"
            ],
            'zambayon': [
                "SAMBAYON", "SAMBAYON PORTOFINO"
            ],
            'oleosa': [
                "ALMENDRADO", "CREMA RUSA", "MARROC"
            ],
            'tortas': [
                "TORTA ALMENDRADO", "TORTA CHOCOTORTA", "TORTA OREO",
                "TORTA PANNACOTTA", "TORTA TRICOLOR"
            ],
            'barras': [
                "BARRA ALMENDRADO", "BARRA CHOCOTORTA", "BARRA OREO",
                "BARRA PANNACOTTA", "BARRA TRICOLOR"
            ],
            'gastronomico': [
                "GASTRO"
            ]
        }

        def clasificar_producto_exacto(nombre_producto):
            """
            Clasifica un producto en UN SOLO grupo (sin duplicación).
            Retorna el nombre del grupo o None.
            """
            nombre_upper = nombre_producto.upper().strip()
            
            # 1. Primero buscar coincidencia EXACTA
            for grupo, nombres in grupos_productos.items():
                for nombre_grupo in nombres:
                    if nombre_upper == nombre_grupo.upper():
                        return grupo
            
            # 2. Luego buscar coincidencia PARCIAL (contiene)
            for grupo, nombres in grupos_productos.items():
                for nombre_grupo in nombres:
                    if nombre_grupo.upper() in nombre_upper:
                        return grupo
            
            return None

        # Clasificar agrupando por producto (1 query, no N+1)
        distribucion_grupos = {grupo: 0 for grupo in grupos_productos.keys()}
        distribucion_grupos['otros'] = 0
        for item in (
            StockBalde.objects
            .filter(balde_plu_q, is_activo=True)
            .values('producto__nombre')
            .annotate(cantidad=Count('id'))
        ):
            grupo = clasificar_producto_exacto(item['producto__nombre'])
            distribucion_grupos[grupo if grupo else 'otros'] += item['cantidad']

        
        # ============================================================
        # 9. ACTIVIDAD POR HORA (Hoy)
        # ============================================================
        
        from django.db.models.functions import TruncHour
        _hora_map = {
            r['hora'].hour: r['total']
            for r in (
                RegistroMovimiento.objects
                .filter(mov_plu_q, timestamp__gte=inicio_dia)
                .annotate(hora=TruncHour('timestamp'))
                .values('hora')
                .annotate(total=Count('id'))
            )
        }
        actividad_horas = [
            {'hora': f'{h:02d}:00', 'movimientos': _hora_map.get(h, 0)}
            for h in range(24)
        ]
        
        # ============================================================
        # 10. ORÍGENES Y DESTINOS MÁS USADOS
        # ============================================================
        
        # Top orígenes (ingresos)
        top_origenes = (
            RegistroMovimiento.objects
            .filter(mov_plu_q, tipo='ingreso', timestamp__gte=hace_30_dias)
            .exclude(origen__isnull=True)
            .exclude(origen='')
            .values('origen')
            .annotate(
                cantidad=Count('id'),
                kg_total=Sum('peso')
            )
            .order_by('-cantidad')[:5]
        )
        
        # Top destinos (retiros) con devoluciones descontadas + desglose por categoría
        _plu_h, _plu_t, _plu_g = _plu_listas_categoria()
        top_destinos_qs = (
            RegistroMovimiento.objects
            .filter(mov_plu_q, tipo='salida', timestamp__gte=hace_30_dias)
            .exclude(boca_salida__isnull=True)
            .exclude(boca_salida='')
            .values('boca_salida')
            .annotate(
                cantidad=Count('id'),
                kg_total=Sum('peso'),
                kg_helados=Sum('peso', filter=Q(producto__plu__in=_plu_h)),
                kg_tortas =Sum('peso', filter=Q(producto__plu__in=_plu_t)),
                kg_gastro =Sum('peso', filter=Q(producto__plu__in=_plu_g)),
            )
            .order_by('-cantidad')[:5]
        )
        # kg devueltos desde cada boca en el mismo período
        dev_map = {
            row['origen']: float(row['kg_dev'] or 0)
            for row in (
                RegistroMovimiento.objects
                .filter(mov_plu_q, tipo='devolucion', timestamp__gte=hace_30_dias)
                .exclude(origen__isnull=True).exclude(origen='')
                .values('origen')
                .annotate(kg_dev=Sum('peso'))
            )
        }
        top_destinos = [
            {
                'boca_salida': d['boca_salida'],
                'cantidad':    d['cantidad'],
                'kg_total':    float(d['kg_total'] or 0),
                'kg_devuelto': round(dev_map.get(d['boca_salida'], 0), 3),
                'kg_neto':     round(float(d['kg_total'] or 0) - dev_map.get(d['boca_salida'], 0), 3),
                'kg_helados':  round(float(d['kg_helados'] or 0), 3),
                'kg_tortas':   round(float(d['kg_tortas'] or 0), 3),
                'kg_gastro':   round(float(d['kg_gastro'] or 0), 3),
            }
            for d in top_destinos_qs
        ]
        
        # ============================================================
        # 11. ESTADÍSTICAS DEL MES
        # ============================================================
        
        stats_mes = RegistroMovimiento.objects.filter(
            mov_plu_q, timestamp__gte=inicio_mes
        ).aggregate(
            total_ingresos=Count('id', filter=Q(tipo='ingreso')),
            total_retiros=Count('id', filter=Q(tipo='salida')),
            kg_ingresados=Sum('peso', filter=Q(tipo='ingreso')),
            kg_retirados=Sum('peso', filter=Q(tipo='salida')),
            peso_promedio=Avg('peso')
        )
        
        # ============================================================
        # 12. ÚLTIMOS MOVIMIENTOS AGRUPADOS
        # ============================================================
        
        ultimos_movimientos = GrupoMovimiento.objects.all().order_by('-fecha')[:10]
        ultimos_movimientos_list = []
        
        for grupo in ultimos_movimientos:
            ultimos_movimientos_list.append({
                'grupo_id': grupo.grupo_id,
                'tipo': grupo.tipo,
                'origen': grupo.origen or '',
                'destino': grupo.destino.nombre if grupo.destino else '',
                'total_peso': float(grupo.total_peso),
                'cantidad_items': grupo.cantidad_items,
                'fecha': grupo.fecha.strftime('%d/%m/%Y %H:%M')
            })
        
        # ============================================================
        # RESPUESTA JSON
        # ============================================================
        
        cats_dash = {
            'helado':       {'baldes': 0, 'kilos': 0.0},
            'barra_torta':  {'baldes': 0, 'kilos': 0.0},
            'gastronomico': {'baldes': 0, 'kilos': 0.0},
        }
        for p in ProductoFijo.objects.annotate(
            cantidad=Count('stockbalde', filter=Q(stockbalde__is_activo=True)),
            kg_total=Sum('stockbalde__peso', filter=Q(stockbalde__is_activo=True)),
        ).values('plu', 'cantidad', 'kg_total'):
            cat = _categorizar_producto(p['plu'])
            cats_dash[cat]['baldes'] += p['cantidad'] or 0
            cats_dash[cat]['kilos']  += float(p['kg_total'] or 0)
        for cat in cats_dash:
            cats_dash[cat]['kilos'] = round(cats_dash[cat]['kilos'], 2)

        response_data = {
            'timestamp': timezone.now().isoformat(),
            'categoria_activa': categoria,
            'resumen_general': {
                'total_baldes': total_baldes,
                'total_kilos': round(float(total_kilos), 2),
                'valor_inventario': valor_inventario,
                'productos_en_stock': productos_en_stock,
                'categorias': cats_dash,
            },
            'movimientos_hoy': {
                'ingresos': movimientos_hoy['ingresos'] or 0,
                'retiros': movimientos_hoy['retiros'] or 0,
                'kg_ingresados': round(float(movimientos_hoy['kg_ingresados'] or 0), 2),
                'kg_retirados': round(float(movimientos_hoy['kg_retirados'] or 0), 2),
                'cambio_ingresos': cambio_ingresos,
                'cambio_retiros': cambio_retiros,
            },
            'top_productos': list(top_productos),
            'productos_bajo_stock': productos_bajo_stock_list,
            'productos_sin_movimiento': list(productos_sin_movimiento),
            'movimientos_7_dias': movimientos_7_dias,
            'movimientos_30_dias': movimientos_30_dias,  # ← NUEVO
            'resumen_30_dias': resumen_30_dias,          # ← NUEVO
            # ✅ Agregar datos custom
            'movimientos_custom': movimientos_custom,
            'resumen_custom': resumen_custom,
            'distribucion_grupos': distribucion_grupos,
            'actividad_horas': actividad_horas,
            'top_origenes': list(top_origenes),
            'top_destinos': list(top_destinos),
            'estadisticas_mes': {
                'total_ingresos': stats_mes['total_ingresos'] or 0,
                'total_retiros': stats_mes['total_retiros'] or 0,
                'kg_ingresados': round(float(stats_mes['kg_ingresados'] or 0), 2),
                'kg_retirados': round(float(stats_mes['kg_retirados'] or 0), 2),
                'peso_promedio': round(float(stats_mes['peso_promedio'] or 0), 2),
            },
            'ultimos_movimientos': ultimos_movimientos_list,
        }
        
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)


# =========================================================
# Auto-updater
# =========================================================

def api_check_update(request):
    from version import APP_VERSION
    from app_inventario.updater import check_for_update
    result = check_for_update(APP_VERSION)
    if result:
        return JsonResponse({"update_available": True, **result})
    return JsonResponse({"update_available": False, "version": APP_VERSION})


@csrf_exempt
def api_apply_update(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body or "{}")
    except Exception:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    download_url = (data.get("download_url") or "").strip()
    if not download_url:
        return JsonResponse({"error": "Falta download_url"}, status=400)
    from app_inventario.updater import download_and_apply_update
    result = download_and_apply_update(download_url)
    if result.get("restart"):
        import threading, os as _os
        threading.Timer(2.0, lambda: _os._exit(0)).start()
    return JsonResponse(result)


# =========================================================
# Conciliación por Boca de Salida
# =========================================================

def conciliacion(request):
    return render(request, "conciliacion.html")


def api_conciliacion_datos(request):
    if request.method != "GET":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    mes_str = (request.GET.get("mes") or "").strip()  # "YYYY-MM"
    if not mes_str:
        hoy = timezone.now().date()
        mes_str = f"{hoy.year}-{hoy.month:02d}"

    try:
        year, month = int(mes_str[:4]), int(mes_str[5:7])
        from datetime import date as _date
        mes_date = _date(year, month, 1)
    except (ValueError, IndexError):
        return JsonResponse({"error": "Formato de mes inválido (YYYY-MM)"}, status=400)

    bocas = BocaSalida.objects.all().order_by("nombre")

    plu_helados, plu_tortas, plu_gastro = _plu_listas_categoria()
    retiros_map = {
        r["boca_salida"]: {
            "kg":         float(r["kg"] or 0),
            "kg_helados": float(r["kg_helados"] or 0),
            "kg_tortas":  float(r["kg_tortas"] or 0),
            "kg_gastro":  float(r["kg_gastro"] or 0),
        }
        for r in (
            RegistroMovimiento.objects
            .filter(tipo="salida", timestamp__year=year, timestamp__month=month)
            .exclude(boca_salida__isnull=True).exclude(boca_salida="")
            .values("boca_salida")
            .annotate(
                kg=Sum("peso"),
                kg_helados=Sum("peso", filter=Q(producto__plu__in=plu_helados)),
                kg_tortas =Sum("peso", filter=Q(producto__plu__in=plu_tortas)),
                kg_gastro =Sum("peso", filter=Q(producto__plu__in=plu_gastro)),
            )
        )
    }

    dev_map = {
        r["origen"]: float(r["kg"] or 0)
        for r in (
            RegistroMovimiento.objects
            .filter(tipo="devolucion", timestamp__year=year, timestamp__month=month)
            .exclude(origen__isnull=True).exclude(origen="")
            .values("origen")
            .annotate(kg=Sum("peso"))
        )
    }

    conc_map = {
        c.boca_id: c
        for c in ConciliacionBoca.objects.filter(mes=mes_date).select_related("boca")
    }

    resultado = []
    for boca in bocas:
        retiro_data  = retiros_map.get(boca.nombre, {})
        kg_recibidos = retiro_data.get("kg", 0.0)
        kg_helados   = retiro_data.get("kg_helados", 0.0)
        kg_tortas    = retiro_data.get("kg_tortas", 0.0)
        kg_gastro    = retiro_data.get("kg_gastro", 0.0)
        kg_devueltos = dev_map.get(boca.nombre, 0.0)
        kg_neto = kg_recibidos - kg_devueltos
        conc = conc_map.get(boca.pk)
        stock_inicial = float(conc.stock_inicial) if conc else 0.0
        kg_vendidos = float(conc.kg_vendidos) if conc else 0.0
        diferencia = stock_inicial + kg_neto - kg_vendidos
        resultado.append({
            "boca_id":      boca.pk,
            "boca_nombre":  boca.nombre,
            "kg_recibidos": round(kg_recibidos, 3),
            "kg_helados":   round(kg_helados, 3),
            "kg_tortas":    round(kg_tortas, 3),
            "kg_gastro":    round(kg_gastro, 3),
            "kg_devueltos": round(kg_devueltos, 3),
            "kg_neto":      round(kg_neto, 3),
            "stock_inicial": round(stock_inicial, 3),
            "kg_vendidos":  round(kg_vendidos, 3),
            "diferencia":   round(diferencia, 3),
        })

    return JsonResponse({"mes": mes_str, "bocas": resultado})


@csrf_exempt
def api_conciliacion_guardar(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    mes_str = (data.get("mes") or "").strip()
    filas = data.get("filas", [])

    try:
        year, month = int(mes_str[:4]), int(mes_str[5:7])
        from datetime import date as _date
        mes_date = _date(year, month, 1)
    except (ValueError, IndexError):
        return JsonResponse({"error": "Mes inválido"}, status=400)

    try:
        with transaction.atomic():
            for fila in filas:
                boca_id = fila.get("boca_id")
                stock_inicial = float(fila.get("stock_inicial") or 0)
                kg_vendidos = float(fila.get("kg_vendidos") or 0)
                try:
                    boca = BocaSalida.objects.get(pk=boca_id)
                except BocaSalida.DoesNotExist:
                    continue
                ConciliacionBoca.objects.update_or_create(
                    boca=boca,
                    mes=mes_date,
                    defaults={"stock_inicial": stock_inicial, "kg_vendidos": kg_vendidos},
                )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"success": True, "message": "Conciliación guardada correctamente."})


def api_conciliacion_exportar(request):
    """GET /api/conciliacion/exportar/?mes=YYYY-MM
    Descarga la conciliación del mes como archivo .xlsx.
    """
    if request.method != "GET":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    mes_str = (request.GET.get("mes") or "").strip()
    if not mes_str:
        hoy = timezone.now().date()
        mes_str = f"{hoy.year}-{hoy.month:02d}"

    try:
        year, month = int(mes_str[:4]), int(mes_str[5:7])
        from datetime import date as _date
        mes_date = _date(year, month, 1)
    except (ValueError, IndexError):
        return JsonResponse({"error": "Formato de mes inválido (YYYY-MM)"}, status=400)

    bocas = BocaSalida.objects.all().order_by("nombre")

    plu_helados, plu_tortas, plu_gastro = _plu_listas_categoria()
    retiros_map = {
        r["boca_salida"]: {
            "kg":         float(r["kg"] or 0),
            "kg_helados": float(r["kg_helados"] or 0),
            "kg_tortas":  float(r["kg_tortas"] or 0),
            "kg_gastro":  float(r["kg_gastro"] or 0),
        }
        for r in (
            RegistroMovimiento.objects
            .filter(tipo="salida", timestamp__year=year, timestamp__month=month)
            .exclude(boca_salida__isnull=True).exclude(boca_salida="")
            .values("boca_salida")
            .annotate(
                kg=Sum("peso"),
                kg_helados=Sum("peso", filter=Q(producto__plu__in=plu_helados)),
                kg_tortas =Sum("peso", filter=Q(producto__plu__in=plu_tortas)),
                kg_gastro =Sum("peso", filter=Q(producto__plu__in=plu_gastro)),
            )
        )
    }

    dev_map = {
        r["origen"]: float(r["kg"] or 0)
        for r in (
            RegistroMovimiento.objects
            .filter(tipo="devolucion", timestamp__year=year, timestamp__month=month)
            .exclude(origen__isnull=True).exclude(origen="")
            .values("origen")
            .annotate(kg=Sum("peso"))
        )
    }

    conc_map = {
        c.boca_id: c
        for c in ConciliacionBoca.objects.filter(mes=mes_date).select_related("boca")
    }

    rows = []
    for boca in bocas:
        retiro_data  = retiros_map.get(boca.nombre, {})
        kg_recibidos = retiro_data.get("kg", 0.0)
        kg_helados   = retiro_data.get("kg_helados", 0.0)
        kg_tortas    = retiro_data.get("kg_tortas", 0.0)
        kg_gastro    = retiro_data.get("kg_gastro", 0.0)
        kg_devueltos = dev_map.get(boca.nombre, 0.0)
        kg_neto = kg_recibidos - kg_devueltos
        conc = conc_map.get(boca.pk)
        stock_inicial = float(conc.stock_inicial) if conc else 0.0
        kg_vendidos = float(conc.kg_vendidos) if conc else 0.0
        diferencia = stock_inicial + kg_neto - kg_vendidos
        rows.append({
            "Boca de Salida":             boca.nombre,
            "Stock Inicial (kg)":         round(stock_inicial, 3),
            "Kg Recibidos":               round(kg_recibidos, 3),
            "↳ Helados (PLU 1-88)":       round(kg_helados, 3),
            "↳ Tortas/Barras (PLU 89-98)": round(kg_tortas, 3),
            "↳ Gastronómico (PLU 100+)":  round(kg_gastro, 3),
            "Kg Devueltos":               round(kg_devueltos, 3),
            "Kg Neto":                    round(kg_neto, 3),
            "Kg Vendidos":                round(kg_vendidos, 3),
            "Diferencia (kg)":            round(diferencia, 3),
        })

    df = pd.DataFrame(rows)

    buffer = BytesIO()
    sheet_name = f"Conc {mes_str}"
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8) + 3
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 28)

    buffer.seek(0)
    filename = f"conciliacion_{mes_str}.xlsx"
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# =========================================================
# Configuración de Precios
# =========================================================

def config_precios(request):
    return render(request, 'config_precios.html')


@csrf_exempt
def api_config_precios(request):
    if request.method == 'GET':
        db_rows = {c.clave: c.valor for c in ConfiguracionSistema.objects.filter(clave__in=_CONFIG_DEFAULTS)}
        configs = {clave: db_rows.get(clave, default) for clave, default in _CONFIG_DEFAULTS.items()}
        return JsonResponse({'config': configs})

    if request.method == 'POST':
        try:
            data = json.loads(request.body or '{}')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        for clave, valor in data.items():
            if clave not in _CONFIG_DEFAULTS:
                continue
            try:
                float(valor)
            except (ValueError, TypeError):
                return JsonResponse({'error': f'Valor inválido para {clave}'}, status=400)
            ConfiguracionSistema.objects.update_or_create(
                clave=clave,
                defaults={'valor': str(int(float(valor))), 'descripcion': ''},
            )
        return JsonResponse({'success': True, 'message': 'Configuración guardada correctamente'})

    return JsonResponse({'error': 'Método no permitido'}, status=405)
